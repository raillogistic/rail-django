"""Django Model Export Functionality

This module provides functionality to export Django model data to Excel or CSV files
through an HTTP endpoint. It supports dynamic model loading, field selection,
filtering, and ordering with GraphQL filter integration.

Features:
- HTTP endpoint for generating downloadable files (JWT protected)
- Support for Excel (.xlsx) and CSV (.csv) formats
- Dynamic model loading by app_name and model_name
- Flexible field selection with nested field access and custom titles
- Advanced filtering using GraphQL filter classes (quick filters, date filters, custom filters)
- Custom ordering support
- Proper error handling and validation
- Default-deny export schema with allowlists and sensitive field blocking
- Formula injection sanitization for CSV/Excel
- Filter/order allowlists with query guardrails
- Optional async exports with job tracking and downloads
- Optional export templates and field formatters

Field Format:
- String format: "field_name" (uses field name as accessor and verbose_name as title)
- Dict format: {"accessor": "field_name", "title": "Custom Title"}

Usage:
    POST /api/v1/export/
    Headers: Authorization: Bearer <jwt_token>
    {
        "app_name": "myapp",
        "model_name": "MyModel",
        "file_extension": "xlsx",
        "filename": "export_data",
        "fields": [
            "title",
            "author.username",
            {"accessor": "slug", "title": "MySlug"}
        ],
        "ordering": ["-created_at"],
        "variables": {
            "status": "active",
            "quick": "search term",
            "published_date_today": true
        }
    }
"""

import csv
import io
import json
import logging
import re
import ipaddress
import tempfile
import threading
import uuid
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Set, Tuple, Union

try:
    from zoneinfo import ZoneInfo
except ImportError:  # pragma: no cover - fallback for older Python
    ZoneInfo = None

from django.apps import apps
from django.conf import settings
from django.core.cache import cache
from django.core.exceptions import FieldDoesNotExist, ImproperlyConfigured
from django.db import models
from django.db.models.fields.related import ForeignKey, ManyToManyField, OneToOneField
from django.db.models.fields.reverse_related import (
    ManyToManyRel,
    ManyToOneRel,
    OneToOneRel,
)
from django.http import (
    FileResponse,
    Http404,
    HttpResponse,
    JsonResponse,
    StreamingHttpResponse,
)
from django.utils import formats, timezone
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from openpyxl.styles import Border, Side

# Import GraphQL filter generator and auth decorators
try:
    from ..generators.filter_inputs import NestedFilterApplicator
except ImportError:
    NestedFilterApplicator = None

try:
    from .auth_decorators import jwt_required
except ImportError:
    jwt_required = None

try:
    from .permissions import OperationType, permission_manager
except ImportError:
    OperationType = None
    permission_manager = None

try:
    from .audit import AuditEventType, log_audit_event
except ImportError:
    AuditEventType = None
    log_audit_event = None

try:
    from ..security.field_permissions import (
        FieldAccessLevel,
        FieldContext,
        field_permission_manager,
    )
except ImportError:
    FieldAccessLevel = None
    FieldContext = None
    field_permission_manager = None

JWT_REQUIRED_AVAILABLE = jwt_required is not None
if jwt_required is None:

    def _missing_jwt_required(view_func):
        raise ImproperlyConfigured(
            "Export endpoints require JWT auth; install auth_decorators to enable."
        )

    _jwt_required = _missing_jwt_required
else:
    _jwt_required = jwt_required

# Optional Excel support
try:
    import openpyxl
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

logger = logging.getLogger(__name__)

DEFAULT_SENSITIVE_FIELDS = [
    "password",
    "passwd",
    "secret",
    "token",
    "access_token",
    "refresh_token",
    "api_key",
    "apikey",
    "private_key",
    "ssh_key",
    "session",
    "ssn",
    "social_security",
    "social_security_number",
    "credit_card",
    "card_number",
    "cvv",
    "cvc",
    "pin",
    "otp",
    "mfa_secret",
]

DEFAULT_ALLOWED_FILTER_LOOKUPS = [
    "exact",
    "iexact",
    "contains",
    "icontains",
    "startswith",
    "istartswith",
    "endswith",
    "iendswith",
    "in",
    "range",
    "isnull",
    "gt",
    "gte",
    "lt",
    "lte",
    "regex",
    "iregex",
]

DEFAULT_ALLOWED_FILTER_TRANSFORMS = [
    "date",
    "year",
    "month",
    "day",
    "week",
    "week_day",
    "quarter",
    "time",
    "hour",
    "minute",
    "second",
]

FORMULA_PREFIXES = ("=", "+", "-", "@")

EXPORT_DEFAULTS = {
    "max_rows": 5000,
    "stream_csv": True,
    "csv_chunk_size": 1000,
    "enforce_streaming_csv": True,
    "excel_write_only": False,  # Use full mode for professional styling
    "excel_auto_width": True,
    "excel_auto_width_max_columns": 50,
    "excel_auto_width_max_rows": 2000,
    "rate_limit": {
        "enable": True,
        "window_seconds": 60,
        "max_requests": 30,
        "trusted_proxies": [],
    },
    "allowed_models": [],
    "allowed_fields": {},  # legacy alias for export_fields
    "export_fields": {},
    "export_exclude": {},
    "sensitive_fields": DEFAULT_SENSITIVE_FIELDS,
    "require_export_fields": False,
    "require_model_permissions": True,
    "require_field_permissions": False,
    "required_permissions": [],
    "allow_callables": False,
    "allow_dunder_access": False,
    "filterable_fields": {},
    "orderable_fields": {},
    "filterable_special_fields": [],
    "allowed_filter_lookups": DEFAULT_ALLOWED_FILTER_LOOKUPS,
    "allowed_filter_transforms": DEFAULT_ALLOWED_FILTER_TRANSFORMS,
    "max_filters": 50,
    "max_or_depth": 3,
    "max_prefetch_depth": 2,
    "sanitize_formulas": True,
    "formula_escape_strategy": "prefix",
    "formula_escape_prefix": "'",
    "field_formatters": {},
    "export_templates": {},
    "async_jobs": {
        "enable": True,
        "backend": "thread",
        "expires_seconds": 3600,
        "storage_dir": None,
        "track_progress": True,
        "progress_update_rows": 500,
    },
}


def _get_export_settings() -> dict[str, Any]:
    """Return merged export settings with defaults applied."""
    export_settings = getattr(settings, "RAIL_DJANGO_EXPORT", None)
    if export_settings is None:
        export_settings = (getattr(settings, "RAIL_DJANGO_GRAPHQL", {}) or {}).get(
            "export_settings", {}
        )

    merged = dict(EXPORT_DEFAULTS)
    if isinstance(export_settings, dict):
        merged.update(export_settings)
        rate_limit_override = export_settings.get("rate_limit")
        rate_limit = dict(EXPORT_DEFAULTS["rate_limit"])
        if isinstance(rate_limit_override, dict):
            rate_limit.update(rate_limit_override)
        merged["rate_limit"] = rate_limit
        async_override = export_settings.get("async_jobs")
        async_jobs = dict(EXPORT_DEFAULTS["async_jobs"])
        if isinstance(async_override, dict):
            async_jobs.update(async_override)
        merged["async_jobs"] = async_jobs

    return merged


def _model_key_candidates(model: type) -> list[str]:
    """Return possible identifiers for a model."""
    return [
        model._meta.label_lower,
        f"{model._meta.app_label}.{model.__name__}",
        model.__name__,
        model._meta.model_name,
    ]


def _normalize_allowed_list(values: Iterable[Any]) -> list[str]:
    """Normalize allowlist values into lowercase strings."""
    return [str(value).strip().lower() for value in values if str(value).strip()]


def _is_model_allowed(model: type, export_settings: dict[str, Any]) -> bool:
    """Check if the model is allowed for export."""
    allowed_models = export_settings.get("allowed_models") or []
    if not allowed_models:
        return True

    allowed = set(_normalize_allowed_list(allowed_models))
    return any(key.lower() in allowed for key in _model_key_candidates(model))


def _get_allowed_fields(model: type, export_settings: dict[str, Any]) -> list[str]:
    """Return the allowed field accessors for a model, if configured."""
    allowed_fields = export_settings.get("allowed_fields") or {}
    if not isinstance(allowed_fields, dict):
        return []

    candidates = {key.lower() for key in _model_key_candidates(model)}
    for key, fields in allowed_fields.items():
        if not key:
            continue
        if str(key).strip().lower() in candidates:
            if isinstance(fields, (list, tuple, set)):
                return [str(field).strip() for field in fields if str(field).strip()]
            return []

    return []


def _normalize_accessor_value(value: str) -> str:
    """Normalize accessor values to dot-notation lowercase."""
    return value.replace("__", ".").strip().lower()


def _normalize_filter_value(value: str) -> str:
    """Normalize filter/order values to __ notation lowercase."""
    return value.replace(".", "__").strip().lower()


def _get_model_scoped_list(model: type, config_value: Any) -> Optional[list[str]]:
    """Return a model-scoped list from a dict keyed by model identifiers."""
    if not isinstance(config_value, dict):
        return None

    candidates = {key.lower() for key in _model_key_candidates(model)}
    for key, values in config_value.items():
        if not key:
            continue
        if str(key).strip().lower() in candidates:
            if isinstance(values, (list, tuple, set)):
                return [str(value).strip() for value in values if str(value).strip()]
            return []

    return []


def _get_model_scoped_dict(model: type, config_value: Any) -> Optional[dict[str, Any]]:
    """Return a model-scoped dict from a dict keyed by model identifiers."""
    if not isinstance(config_value, dict):
        return None

    candidates = {key.lower() for key in _model_key_candidates(model)}
    for key, values in config_value.items():
        if not key:
            continue
        if str(key).strip().lower() in candidates and isinstance(values, dict):
            return values

    return None


def _get_export_fields(model: type, export_settings: dict[str, Any]) -> list[str]:
    """Return explicit export field allowlist (full-path)."""
    export_fields = export_settings.get("export_fields")
    if export_fields is None:
        export_fields = export_settings.get("allowed_fields")
    scoped = _get_model_scoped_list(model, export_fields)
    if scoped is None:
        return []
    return [_normalize_accessor_value(value) for value in scoped]


def _get_export_exclude(model: type, export_settings: dict[str, Any]) -> list[str]:
    """Return explicit export field denylist (full-path)."""
    scoped = _get_model_scoped_list(model, export_settings.get("export_exclude"))
    if scoped is None:
        return []
    return [_normalize_accessor_value(value) for value in scoped]


def _get_filterable_fields(
    model: type, export_settings: dict[str, Any], export_fields: list[str]
) -> list[str]:
    """Return filterable fields (full-path) in __ notation."""
    scoped = _get_model_scoped_list(model, export_settings.get("filterable_fields"))
    if scoped is None or not scoped:
        scoped = export_fields
    return [_normalize_filter_value(value) for value in scoped]


def _get_orderable_fields(
    model: type, export_settings: dict[str, Any], export_fields: list[str]
) -> list[str]:
    """Return orderable fields (full-path) in __ notation."""
    scoped = _get_model_scoped_list(model, export_settings.get("orderable_fields"))
    if scoped is None or not scoped:
        scoped = export_fields
    return [_normalize_filter_value(value) for value in scoped]


def _get_field_formatters(
    model: type, export_settings: dict[str, Any]
) -> dict[str, Any]:
    """Return field formatter mappings for a model."""
    formatters = export_settings.get("field_formatters") or {}
    scoped = _get_model_scoped_dict(model, formatters)
    if scoped is not None:
        return {
            _normalize_accessor_value(key): value
            for key, value in scoped.items()
            if isinstance(key, str)
        }
    if isinstance(formatters, dict):
        return {
            _normalize_accessor_value(key): value
            for key, value in formatters.items()
            if isinstance(key, str)
        }
    return {}


def _get_export_templates(export_settings: dict[str, Any]) -> dict[str, Any]:
    """Return configured export templates."""
    templates = export_settings.get("export_templates") or {}
    if not isinstance(templates, dict):
        return {}
    return templates


def _sanitize_filename(filename: str) -> str:
    """Sanitize filenames for safe Content-Disposition and filesystem usage."""
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", filename).strip("._")
    return cleaned or "export"


def _export_job_cache_key(job_id: str) -> str:
    return f"rail:export_job:{job_id}"


def _export_job_payload_key(job_id: str) -> str:
    return f"rail:export_job_payload:{job_id}"


def _get_export_storage_dir(export_settings: dict[str, Any]) -> Path:
    async_settings = export_settings.get("async_jobs") or {}
    storage_dir = async_settings.get("storage_dir")
    if storage_dir:
        path = Path(str(storage_dir))
    elif getattr(settings, "MEDIA_ROOT", None):
        path = Path(settings.MEDIA_ROOT) / "rail_exports"
    else:
        path = Path(tempfile.gettempdir()) / "rail_exports"
    path.mkdir(parents=True, exist_ok=True)
    return path


def _get_export_job(job_id: str) -> Optional[dict[str, Any]]:
    return cache.get(_export_job_cache_key(job_id))


def _set_export_job(job_id: str, job: dict[str, Any], *, timeout: int) -> None:
    cache.set(_export_job_cache_key(job_id), job, timeout=timeout)


def _update_export_job(
    job_id: str, updates: dict[str, Any], *, timeout: int
) -> Optional[dict[str, Any]]:
    job = _get_export_job(job_id)
    if not job:
        return None
    job.update(updates)
    _set_export_job(job_id, job, timeout=timeout)
    return job


def _delete_export_job(job_id: str) -> None:
    cache.delete(_export_job_cache_key(job_id))
    cache.delete(_export_job_payload_key(job_id))


def _run_export_job(job_id: str) -> None:
    """Execute an async export job and update cache state."""
    job = _get_export_job(job_id)
    if not job:
        return

    export_settings = _get_export_settings()
    async_settings = export_settings.get("async_jobs") or {}
    timeout = int(async_settings.get("expires_seconds", 3600))
    payload = cache.get(_export_job_payload_key(job_id))
    if not payload:
        _update_export_job(
            job_id,
            {"status": "failed", "error": "Missing job payload"},
            timeout=timeout,
        )
        return

    _update_export_job(
        job_id,
        {"status": "running", "started_at": timezone.now().isoformat()},
        timeout=timeout,
    )

    processed_rows = 0
    total_rows = None

    def progress_callback(count: int) -> None:
        nonlocal processed_rows
        processed_rows = count
        _update_export_job(
            job_id,
            {"processed_rows": processed_rows},
            timeout=timeout,
        )

    try:
        exporter = ModelExporter(
            payload["app_name"],
            payload["model_name"],
            export_settings=export_settings,
        )
        parsed_fields = payload.get("parsed_fields") or exporter.validate_fields(
            payload["fields"], export_settings=export_settings
        )
        ordering = payload.get("ordering")
        variables = payload.get("variables") or {}
        max_rows = payload.get("max_rows")
        file_extension = payload["file_extension"]
        filename = payload["filename"]
        if file_extension not in {"csv", "xlsx"}:
            raise ExportError("Unsupported export format")

        storage_dir = _get_export_storage_dir(export_settings)
        file_path = storage_dir / f"{job_id}.{file_extension}"

        if async_settings.get("track_progress", True):
            try:
                total_rows = exporter.get_queryset(
                    variables,
                    ordering,
                    fields=[field["accessor"] for field in parsed_fields],
                    max_rows=max_rows,
                ).count()
            except Exception:
                total_rows = None
            _update_export_job(
                job_id,
                {"total_rows": total_rows},
                timeout=timeout,
            )

        if file_extension == "csv":
            with open(file_path, "w", encoding="utf-8", newline="") as handle:
                exporter.export_to_csv(
                    payload["fields"],
                    variables,
                    ordering,
                    max_rows=max_rows,
                    parsed_fields=parsed_fields,
                    output=handle,
                    progress_callback=progress_callback,
                )
            content_type = "text/csv; charset=utf-8"
        else:
            with open(file_path, "wb") as handle:
                exporter.export_to_excel(
                    payload["fields"],
                    variables,
                    ordering,
                    max_rows=max_rows,
                    parsed_fields=parsed_fields,
                    output=handle,
                    progress_callback=progress_callback,
                )
            content_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        if total_rows is not None and total_rows > processed_rows:
            processed_rows = total_rows
        _update_export_job(
            job_id,
            {
                "status": "completed",
                "completed_at": timezone.now().isoformat(),
                "file_path": str(file_path),
                "content_type": content_type,
                "filename": filename,
                "processed_rows": processed_rows,
            },
            timeout=timeout,
        )

    except Exception as exc:
        _update_export_job(
            job_id,
            {
                "status": "failed",
                "error": str(exc),
                "completed_at": timezone.now().isoformat(),
            },
            timeout=timeout,
        )


def _parse_iso_datetime(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(value)
    except ValueError:
        return None
    if timezone.is_naive(parsed):
        return timezone.make_aware(parsed, timezone.get_default_timezone())
    return parsed


def _cleanup_export_job_files(job: dict[str, Any]) -> None:
    file_path = job.get("file_path")
    if not file_path:
        return
    try:
        path = Path(file_path)
        if path.exists():
            path.unlink()
    except Exception:
        return


def _job_access_allowed(request: Any, job: dict[str, Any]) -> bool:
    user = getattr(request, "user", None)
    if not user or not getattr(user, "is_authenticated", False):
        return False
    if getattr(user, "is_superuser", False):
        return True
    owner_id = job.get("owner_id")
    return owner_id is not None and str(owner_id) == str(user.id)


try:
    from celery import shared_task
except Exception:
    shared_task = None

if shared_task:

    @shared_task(name="rail_django.export_job")
    def export_job_task(job_id: str) -> None:
        _run_export_job(job_id)
else:
    export_job_task = None


class ExportError(Exception):
    """Custom exception for export-related errors."""

    pass


class ModelExporter:
    """
    Handles the export of Django model data to various formats.

    This class provides methods to dynamically load models, apply GraphQL filters,
    extract field data with flexible field format support, and generate export files.

    Features:
    - Dynamic model loading from app and model names
    - GraphQL filter integration for advanced filtering
    - Flexible field format support (string or dict with accessor/title)
    - Nested field access with proper error handling
    - Property access on model instances (explicit allowlist)
    - Many-to-many field handling
    """

    def __init__(
        self,
        app_name: str,
        model_name: str,
        *,
        export_settings: Optional[dict[str, Any]] = None,
        schema_name: Optional[str] = None,
    ):
        """
        Initialize the exporter with model information and GraphQL filter generator.

        Args:
            app_name: Name of the Django app containing the model
            model_name: Name of the Django model to export
            schema_name: Schema name for multi-schema filter support (defaults to app_name or "default")

        Raises:
            ExportError: If the model cannot be found
        """
        self.app_name = app_name
        self.model_name = model_name
        self.export_settings = export_settings or _get_export_settings()
        self.schema_name = schema_name or app_name or "default"
        self.model = self._load_model()
        self.logger = logging.getLogger(__name__)
        self.allow_callables = bool(self.export_settings.get("allow_callables", False))
        self.allow_dunder_access = bool(
            self.export_settings.get("allow_dunder_access", False)
        )
        self.sanitize_formulas = bool(
            self.export_settings.get("sanitize_formulas", True)
        )
        self.formula_escape_strategy = str(
            self.export_settings.get("formula_escape_strategy", "prefix")
        ).lower()
        self.formula_escape_prefix = str(
            self.export_settings.get("formula_escape_prefix", "'")
        )
        self.sensitive_fields = [
            value.lower()
            for value in (self.export_settings.get("sensitive_fields") or [])
            if str(value).strip()
        ]
        self.export_fields = _get_export_fields(self.model, self.export_settings)
        self.export_exclude = _get_export_exclude(self.model, self.export_settings)
        self.max_prefetch_depth = self._normalize_max_depth(
            self.export_settings.get("max_prefetch_depth")
        )
        self.filterable_fields = _get_filterable_fields(
            self.model, self.export_settings, self.export_fields
        )
        self.orderable_fields = _get_orderable_fields(
            self.model, self.export_settings, self.export_fields
        )
        self.filterable_special_fields = [
            value.strip().lower()
            for value in (self.export_settings.get("filterable_special_fields") or [])
            if str(value).strip()
        ]
        self.allowed_filter_lookups = [
            value.strip().lower()
            for value in (self.export_settings.get("allowed_filter_lookups") or [])
            if str(value).strip()
        ]
        self.allowed_filter_transforms = [
            value.strip().lower()
            for value in (self.export_settings.get("allowed_filter_transforms") or [])
            if str(value).strip()
        ]
        self.field_formatters = _get_field_formatters(self.model, self.export_settings)

        # Initialize GraphQL filter applicator if available (singleton pattern)
        self.nested_filter_applicator = None
        if NestedFilterApplicator:
            try:
                from ..generators.filter_inputs import get_nested_filter_applicator

                self.nested_filter_applicator = get_nested_filter_applicator(
                    self.schema_name
                )
                self.logger.info("Nested filter applicator initialized successfully")
            except Exception as e:
                self.logger.warning(
                    f"Failed to initialize nested filter applicator: {e}"
                )

    def _load_model(self) -> models.Model:
        """
        Load the Django model dynamically.

        Returns:
            The Django model class

        Raises:
            ExportError: If the model cannot be found
        """
        try:
            return apps.get_model(self.app_name, self.model_name)
        except LookupError as e:
            raise ExportError(
                f"Model '{self.model_name}' not found in app '{self.app_name}': {e}"
            )

    def _normalize_ordering(
        self, ordering: Optional[Union[str, list[str]]]
    ) -> list[str]:
        """Normalize and validate ordering input into a list of field expressions."""
        if not ordering:
            return []
        if isinstance(ordering, str):
            items = [ordering]
        elif isinstance(ordering, (list, tuple)):
            items = [item for item in ordering if isinstance(item, str) and item]
        else:
            return []

        normalized: list[str] = []
        invalid: list[str] = []
        for item in items:
            desc = item.startswith("-")
            field_name = item[1:] if desc else item
            if not self._is_orderable(field_name):
                invalid.append(item)
                continue
            normalized.append(item)

        if invalid:
            raise ExportError(
                "Ordering not allowed: " + ", ".join(sorted(set(invalid)))
            )

        return normalized

    def _normalize_max_depth(self, value: Any) -> Optional[int]:
        """Normalize a depth limit to a positive int or None."""
        if value is None:
            return None
        try:
            value = int(value)
        except (TypeError, ValueError):
            return None
        if value <= 0:
            return None
        return value

    def _is_orderable(self, field_name: str) -> bool:
        """Check whether a field is allowed for ordering."""
        if not field_name:
            return False
        parts = field_name.replace("__", ".").split(".")
        if any(part.startswith("_") for part in parts):
            return False
        # If no orderable_fields configured, allow all (permissive default)
        if not self.orderable_fields:
            return True
        normalized = _normalize_filter_value(field_name)
        return normalized in self.orderable_fields

    def _resolve_model_field(
        self, model_class: type, field_name: str
    ) -> Optional[models.Field]:
        """Resolve a field or reverse relation on a model."""
        try:
            return model_class._meta.get_field(field_name)
        except FieldDoesNotExist:
            pass

        related_objects = getattr(model_class._meta, "related_objects", [])
        for relation in related_objects:
            if relation.get_accessor_name() == field_name:
                return relation

        return None

    def _collect_related_paths(self, accessors: Iterable[str]) -> dict[str, list[str]]:
        """Collect select_related and prefetch_related paths for accessors."""
        select_related: set[str] = set()
        prefetch_related: set[str] = set()

        for accessor in accessors:
            if not accessor:
                continue
            parts = accessor.split(".")
            current_model = self.model
            path_parts: list[str] = []
            prefetch_mode = False
            relation_depth = 0

            for part in parts:
                if part.endswith("()"):
                    break
                field = self._resolve_model_field(current_model, part)
                if not field:
                    break

                path_parts.append(part)

                if isinstance(field, (ForeignKey, OneToOneField)):
                    relation_depth += 1
                    if (
                        self.max_prefetch_depth
                        and relation_depth > self.max_prefetch_depth
                    ):
                        raise ExportError(
                            f"Max prefetch depth exceeded for accessor '{accessor}'"
                        )
                    if prefetch_mode:
                        prefetch_related.add("__".join(path_parts))
                    else:
                        select_related.add("__".join(path_parts))
                    current_model = field.related_model
                    continue

                if isinstance(
                    field, (ManyToManyField, ManyToOneRel, ManyToManyRel, OneToOneRel)
                ):
                    relation_depth += 1
                    if (
                        self.max_prefetch_depth
                        and relation_depth > self.max_prefetch_depth
                    ):
                        raise ExportError(
                            f"Max prefetch depth exceeded for accessor '{accessor}'"
                        )
                    prefetch_mode = True
                    prefetch_related.add("__".join(path_parts))
                    related_model = getattr(field, "related_model", None)
                    if related_model:
                        current_model = related_model
                    continue

                break

        return {
            "select_related": sorted(select_related),
            "prefetch_related": sorted(prefetch_related),
        }

    def _apply_related_optimizations(
        self, queryset: models.QuerySet, accessors: Iterable[str]
    ) -> models.QuerySet:
        """Apply select_related/prefetch_related based on accessors."""
        related_paths = self._collect_related_paths(accessors)
        if related_paths["select_related"]:
            queryset = queryset.select_related(*related_paths["select_related"])
        if related_paths["prefetch_related"]:
            queryset = queryset.prefetch_related(*related_paths["prefetch_related"])
        return queryset

    def _has_field_access(self, user: Any, accessor: str) -> bool:
        """Check field-level access permissions if configured."""
        if not field_permission_manager or not FieldContext or not FieldAccessLevel:
            return True
        if not user or not getattr(user, "is_authenticated", False):
            return False
        if getattr(user, "is_superuser", False):
            return True

        current_model = self.model
        for part in accessor.split("."):
            if part.endswith("()"):
                return False
            if not self.allow_dunder_access and part.startswith("_"):
                return False
            context = FieldContext(
                user=user,
                field_name=part,
                operation_type="read",
                model_class=current_model,
            )
            access_level = field_permission_manager.get_field_access_level(context)
            if access_level == FieldAccessLevel.NONE:
                return False

            field = self._resolve_model_field(current_model, part)
            related_model = getattr(field, "related_model", None) if field else None
            if related_model:
                current_model = related_model
            else:
                break

        return True

    def _validate_accessor(
        self,
        accessor: str,
        export_fields: list[str],
        export_exclude: list[str],
        sensitive_fields: list[str],
        require_export_fields: bool,
    ) -> Optional[str]:
        """Validate accessor syntax and model traversal."""
        if "__" in accessor:
            return "Accessors must use dot notation"
        normalized = _normalize_accessor_value(accessor)
        if not normalized:
            return "Empty accessor"
        if normalized in export_exclude:
            return "Accessor is explicitly excluded"
        if export_fields and normalized not in export_fields:
            return "Accessor is not allowlisted"

        parts = accessor.split(".")
        for part in parts:
            if not part:
                return "Accessor contains empty path segments"
            if not self.allow_dunder_access and part.startswith("_"):
                return "Accessor uses a private field"
            if part.endswith("()") and not self.allow_callables:
                return "Callable accessors are disabled"
            if part.endswith("()") and part[:-2].startswith("_"):
                return "Accessor uses a private field"

        if sensitive_fields:
            for part in parts:
                if part.lower() in sensitive_fields:
                    return "Accessor matches a sensitive field"

        current_model = self.model
        relation_depth = 0

        for index, part in enumerate(parts):
            is_last = index == len(parts) - 1
            part_name = part[:-2] if part.endswith("()") else part
            if part.endswith("()") and not is_last:
                return "Callable accessors cannot be chained"
            field = self._resolve_model_field(current_model, part_name)

            if field is None:
                if not is_last:
                    return "Accessor cannot traverse non-relational field"
                attr = getattr(current_model, part_name, None)
                if callable(attr) and not self.allow_callables:
                    return "Callable accessors are disabled"
                return None

            is_relation = isinstance(
                field,
                (
                    ForeignKey,
                    OneToOneField,
                    ManyToManyField,
                    ManyToOneRel,
                    ManyToManyRel,
                    OneToOneRel,
                ),
            )
            if is_relation:
                relation_depth += 1
                if self.max_prefetch_depth and relation_depth > self.max_prefetch_depth:
                    return "Accessor exceeds max relationship depth"
                related_model = getattr(field, "related_model", None)
                if related_model and not is_last:
                    current_model = related_model
                    continue
                if not is_last:
                    return "Accessor cannot traverse non-relational field"
                return None

            if not is_last:
                return "Accessor cannot traverse non-relational field"

        return None

    def validate_fields(
        self,
        fields: list[Union[str, dict[str, str]]],
        *,
        user: Optional[Any] = None,
        export_settings: Optional[dict[str, Any]] = None,
    ) -> list[dict[str, str]]:
        """Validate and normalize fields based on allowlist and permissions."""
        export_settings = export_settings or self.export_settings
        export_fields = _get_export_fields(self.model, export_settings)
        export_exclude = _get_export_exclude(self.model, export_settings)
        sensitive_fields = [
            value.lower()
            for value in (export_settings.get("sensitive_fields") or [])
            if str(value).strip()
        ]
        require_export_fields = bool(export_settings.get("require_export_fields", True))
        require_field_permissions = bool(
            export_settings.get("require_field_permissions", True)
        )

        if require_export_fields and not export_fields:
            raise ExportError(
                f"Export denied: schema is not configured for model {self.model._meta.label}"
            )

        parsed_fields: list[dict[str, str]] = []
        denied_fields: list[str] = []

        for field_config in fields:
            parsed_field = self.parse_field_config(field_config)
            accessor = parsed_field.get("accessor", "").strip()
            if not accessor:
                denied_fields.append("<empty>")
                continue

            error = self._validate_accessor(
                accessor,
                export_fields,
                export_exclude,
                sensitive_fields,
                require_export_fields,
            )
            if error:
                denied_fields.append(accessor)
                continue

            if (
                user is not None
                and require_field_permissions
                and not self._has_field_access(user, accessor)
            ):
                denied_fields.append(accessor)
                continue

            parsed_fields.append(parsed_field)

        if denied_fields:
            raise ExportError(
                "Export denied for fields: " + ", ".join(sorted(set(denied_fields)))
            )

        if not parsed_fields:
            raise ExportError("No exportable fields were provided")

        return parsed_fields

    def _analyze_filter_tree(
        self, filter_input: Any, *, current_or_depth: int = 0
    ) -> tuple[int, int]:
        """Return (filter_count, max_or_depth) for a filter tree."""
        if not filter_input:
            return 0, current_or_depth

        if isinstance(filter_input, list):
            total = 0
            max_or_depth = current_or_depth
            for item in filter_input:
                count, depth = self._analyze_filter_tree(
                    item, current_or_depth=current_or_depth
                )
                total += count
                max_or_depth = max(max_or_depth, depth)
            return total, max_or_depth

        if not isinstance(filter_input, dict):
            return 0, current_or_depth

        total_filters = 0
        max_or_depth = current_or_depth

        for key, value in filter_input.items():
            if key == "AND":
                count, depth = self._analyze_filter_tree(
                    value, current_or_depth=current_or_depth
                )
                total_filters += count
                max_or_depth = max(max_or_depth, depth)
                continue
            if key == "OR":
                count, depth = self._analyze_filter_tree(
                    value, current_or_depth=current_or_depth + 1
                )
                total_filters += count
                max_or_depth = max(max_or_depth, depth)
                continue
            if key == "NOT":
                count, depth = self._analyze_filter_tree(
                    value, current_or_depth=current_or_depth
                )
                total_filters += count
                max_or_depth = max(max_or_depth, depth)
                continue

            total_filters += 1

        return total_filters, max_or_depth

    def _iter_filter_keys(self, filter_input: Any) -> Iterable[str]:
        """Yield filter keys from a filter tree."""
        if not filter_input:
            return

        if isinstance(filter_input, list):
            for item in filter_input:
                yield from self._iter_filter_keys(item)
            return

        if not isinstance(filter_input, dict):
            return

        for key, value in filter_input.items():
            if key in {"AND", "OR", "NOT"}:
                yield from self._iter_filter_keys(value)
                continue
            yield key

    def _is_filter_key_allowed(self, key: str) -> bool:
        """Check whether a filter key is allowlisted."""
        if not key:
            return False
        normalized = _normalize_filter_value(key)
        parts_for_private = normalized.split("__")
        if any(part.startswith("_") for part in parts_for_private):
            return False
        if normalized in self.filterable_special_fields:
            return True
        # If no filterable_fields configured, allow all (permissive default)
        if not self.filterable_fields:
            return True

        parts = normalized.split("__")
        if parts[-1] in self.allowed_filter_lookups:
            base_parts = parts[:-1]
            while base_parts and base_parts[-1] in self.allowed_filter_transforms:
                base_parts = base_parts[:-1]
            if not base_parts:
                return False
            base = "__".join(base_parts)
            return base in self.filterable_fields

        if parts[-1] in self.allowed_filter_transforms:
            base_parts = parts
            while base_parts and base_parts[-1] in self.allowed_filter_transforms:
                base_parts = base_parts[:-1]
            if not base_parts:
                return False
            base = "__".join(base_parts)
            return base in self.filterable_fields

        return normalized in self.filterable_fields

    def validate_filter_input(
        self,
        where_input: Optional[dict[str, Any]] = None,
        *,
        export_settings: Optional[dict[str, Any]] = None,
    ) -> None:
        """
        Validate filter input against export allowlists and complexity limits.

        This method standardizes on the "where" key format and uses validate_filter_complexity
        from filter_inputs for depth and clause count validation.

        Args:
            where_input: The where filter dictionary (the contents of variables["where"])
            export_settings: Optional export settings override

        Raises:
            ExportError: If filters violate complexity or allowlist rules
        """
        if not where_input:
            return
        if not isinstance(where_input, dict):
            raise ExportError("where must be an object")

        export_settings = export_settings or self.export_settings

        # Import and use complexity validation from filter_inputs
        try:
            from ..generators.filter_inputs import validate_filter_complexity

            max_depth = export_settings.get("max_or_depth") or 10
            max_clauses = export_settings.get("max_filters") or 50
            try:
                max_depth = int(max_depth)
            except (TypeError, ValueError):
                max_depth = 10
            try:
                max_clauses = int(max_clauses)
            except (TypeError, ValueError):
                max_clauses = 50
            if max_depth <= 0:
                max_depth = 10
            if max_clauses <= 0:
                max_clauses = 50
            validate_filter_complexity(
                where_input, max_depth=max_depth, max_clauses=max_clauses
            )
        except ImportError:
            # Fall back to existing validation if filter_inputs unavailable
            max_filters = export_settings.get("max_filters", None)
            max_or_depth = export_settings.get("max_or_depth", None)
            total_filters, max_depth_found = self._analyze_filter_tree(where_input)

            if max_filters is not None:
                try:
                    max_filters = int(max_filters)
                except (TypeError, ValueError):
                    max_filters = None
            if max_filters is not None and max_filters <= 0:
                max_filters = None
            if max_filters is not None and total_filters > max_filters:
                raise ExportError("Too many filters were provided")

            if max_or_depth is not None:
                try:
                    max_or_depth = int(max_or_depth)
                except (TypeError, ValueError):
                    max_or_depth = None
            if max_or_depth is not None and max_or_depth <= 0:
                max_or_depth = None
            if max_or_depth is not None and max_depth_found > max_or_depth:
                raise ExportError("Filter OR depth exceeds limit")
        except Exception as e:
            raise ExportError(f"Filter complexity error: {e}")

        # Validate against export-specific field allowlists
        invalid_keys = [
            key
            for key in self._iter_filter_keys(where_input)
            if not self._is_filter_key_allowed(key)
        ]
        if invalid_keys:
            raise ExportError(
                "Filters not allowed: " + ", ".join(sorted(set(invalid_keys)))
            )

    def validate_filters(
        self,
        variables: Optional[dict[str, Any]] = None,
        *,
        export_settings: Optional[dict[str, Any]] = None,
    ) -> None:
        """
        Validate filter input against allowlists and guardrails.

        .. deprecated::
            Use :meth:`validate_filter_input` instead. This method is kept for
            backward compatibility and will be removed in a future version.
        """
        import warnings

        warnings.warn(
            "validate_filters() is deprecated; use validate_filter_input() instead",
            DeprecationWarning,
            stacklevel=2,
        )
        if not variables:
            return

        export_settings = export_settings or self.export_settings
        # Support both "where" and "filters" keys for backward compatibility
        filter_input = variables.get("where") or variables.get("filters", variables)
        self.validate_filter_input(filter_input, export_settings=export_settings)

    def get_queryset(
        self,
        variables: Optional[dict[str, Any]] = None,
        ordering: Optional[Union[str, list[str]]] = None,
        fields: Optional[Iterable[str]] = None,
        max_rows: Optional[int] = None,
        *,
        presets: Optional[List[str]] = None,
        skip_validation: bool = False,
        distinct_on: Optional[List[str]] = None,
    ) -> models.QuerySet:
        """
        Get the filtered and ordered queryset using GraphQL filters.

        Args:
            variables: Dictionary of filter kwargs (expects {"where": {...}})
            ordering: Django ORM ordering expression(s)
            fields: Field accessors for select_related/prefetch_related optimization
            max_rows: Optional max rows cap
            presets: Optional list of preset names to apply from GraphQLMeta.filter_presets
            skip_validation: If True, skip filter validation (use when already validated)
            distinct_on: Optional list of field names for DISTINCT ON (PostgreSQL only)

        Returns:
            Filtered and ordered queryset

        Raises:
            ExportError: If filtering or ordering fails
        """
        try:
            queryset = self.model.objects.all()

            # Apply GraphQL filters
            if variables:
                if not skip_validation:
                    where_input = variables.get("where", variables)
                    self.validate_filter_input(where_input)
                queryset = self.apply_graphql_filters(
                    queryset, variables, presets=presets
                )

            # Apply ordering (must come before distinct for PostgreSQL DISTINCT ON)
            ordering_fields = self._normalize_ordering(ordering)
            if ordering_fields:
                queryset = queryset.order_by(*ordering_fields)

            # Apply DISTINCT ON if specified (PostgreSQL only)
            if distinct_on:
                distinct_fields = [
                    f.replace(".", "__")
                    for f in distinct_on
                    if isinstance(f, str) and f
                ]
                if distinct_fields:
                    queryset = queryset.distinct(*distinct_fields)

            # Apply relation optimizations based on requested fields
            if fields:
                queryset = self._apply_related_optimizations(queryset, fields)

            if max_rows is not None and max_rows > 0:
                queryset = queryset[:max_rows]

            return queryset

        except Exception as e:
            raise ExportError(f"Error building queryset: {e}")

    def apply_graphql_filters(
        self,
        queryset: models.QuerySet,
        variables: dict[str, Any],
        *,
        presets: Optional[List[str]] = None,
    ) -> models.QuerySet:
        """
        Apply GraphQL filters to the queryset using the nested filter applicator.

        This method standardizes on the "where" key format and supports filter presets.
        Errors are raised immediately instead of silently falling back to basic filtering.

        Args:
            queryset: Django QuerySet to filter
            variables: Filter parameters from the request (expects {"where": {...}})
            presets: Optional list of preset names to apply from GraphQLMeta.filter_presets

        Returns:
            Filtered QuerySet

        Raises:
            ExportError: If nested filter applicator is unavailable or filtering fails
        """
        if not variables:
            return queryset

        # Standardize on "where" key
        where_input = variables.get("where", variables)
        if not where_input:
            return queryset
        if not isinstance(where_input, dict):
            raise ExportError("where must be an object")

        if not self.nested_filter_applicator:
            raise ExportError(
                "Nested filter applicator not available. Ensure filter_inputs module is accessible."
            )

        try:
            # Apply presets if provided
            if presets:
                where_input = self.nested_filter_applicator.apply_presets(
                    where_input, presets, self.model
                )
            return self.nested_filter_applicator.apply_where_filter(
                queryset, where_input, self.model
            )
        except Exception as e:
            # NO silent fallback - fail clearly
            raise ExportError(f"Filter application failed: {e}")

    def get_field_value(self, instance: models.Model, accessor: str) -> Any:
        """
        Get field value from model instance using accessor path.

        Supports:
        - Simple fields: 'title'
        - Nested fields: 'author.username'
        - Many-to-many fields: 'tags' (returns comma-separated list)

        Args:
            instance: Model instance
            accessor: Dot-separated path to the field/attribute

        Returns:
            The field value, properly formatted
        """
        try:
            # Split accessor by dots for nested access
            parts = accessor.split(".")
            value = instance

            for part in parts:
                if value is None:
                    return None

                # Handle method calls (if part ends with parentheses)
                if part.endswith("()"):
                    if not self.allow_callables:
                        return None
                    method_name = part[:-2]
                    if not self.allow_dunder_access and method_name.startswith("_"):
                        return None
                    method = getattr(value, method_name, None)
                    if callable(method):
                        value = method()
                    else:
                        return None
                else:
                    # Regular attribute access
                    if not self.allow_dunder_access and part.startswith("_"):
                        return None
                    if not hasattr(value, part):
                        return None
                    attr = getattr(value, part)
                    if callable(attr):
                        if not self.allow_callables:
                            return None
                        try:
                            value = attr()
                        except Exception as e:
                            self.logger.debug(f"Callable access failed for {part}: {e}")
                            return None
                    else:
                        value = attr

            # Handle many-to-many relationships
            if hasattr(value, "all"):
                try:
                    items = list(value.all())
                    if items:
                        value = ", ".join(str(item) for item in items)
                    else:
                        value = ""
                except Exception:
                    pass

            value = self._apply_field_formatter(value, accessor)
            return self._format_value(value)

        except Exception as e:
            self.logger.warning(
                f"Error accessing field '{accessor}' on {instance}: {e}"
            )
            return None

    def _format_value(self, value: Any) -> Any:
        """
        Format value for export based on its type.

        Args:
            value: The value to format

        Returns:
            Formatted value suitable for export
        """
        if value is None:
            return ""
        if isinstance(value, bool):
            formatted: Any = "Yes" if value else "No"
        elif isinstance(value, (datetime, date)):
            if isinstance(value, datetime):
                # Convert timezone-aware datetime to local time
                if timezone.is_aware(value):
                    value = timezone.localtime(value)
                formatted = value.strftime("%Y-%m-%d %H:%M:%S")
            else:
                formatted = value.strftime("%Y-%m-%d")
        elif isinstance(value, Decimal):
            formatted = float(value)
        elif isinstance(value, models.Model):
            # For related objects, return string representation
            formatted = str(value)
        elif hasattr(value, "all"):
            # For many-to-many fields, join related objects
            formatted = ", ".join(str(item) for item in value.all())
        else:
            formatted = str(value)

        if isinstance(formatted, str):
            return self._apply_formula_sanitizer(formatted)
        return formatted

    def _apply_formula_sanitizer(self, value: str) -> str:
        """Escape values that could be interpreted as formulas by spreadsheet tools."""
        if not self.sanitize_formulas or not value:
            return value
        stripped = value.lstrip()
        if stripped and stripped[0] in FORMULA_PREFIXES:
            if self.formula_escape_strategy == "prefix":
                return f"{self.formula_escape_prefix}{value}"
            if self.formula_escape_strategy == "tab":
                return f"\t{value}"
        return value

    def _apply_field_formatter(self, value: Any, accessor: str) -> Any:
        """Apply per-field formatting or masking."""
        formatter = self.field_formatters.get(_normalize_accessor_value(accessor))
        if not formatter:
            return value
        if isinstance(formatter, str):
            formatter = {"type": formatter}

        formatter_type = str(formatter.get("type", "")).lower()
        if formatter_type == "redact":
            return formatter.get("value", "[REDACTED]")
        if formatter_type == "mask":
            raw = "" if value is None else str(value)
            show_last = int(formatter.get("show_last", 4))
            mask_char = str(formatter.get("mask_char", "*"))
            if show_last <= 0:
                return mask_char * len(raw)
            masked = mask_char * max(len(raw) - show_last, 0)
            return f"{masked}{raw[-show_last:]}" if raw else raw
        if formatter_type in {"datetime", "date"}:
            if not isinstance(value, (datetime, date)):
                return value
            tz_name = formatter.get("timezone")
            if isinstance(value, datetime):
                if timezone.is_naive(value):
                    value = timezone.make_aware(value, timezone.get_default_timezone())
                if tz_name and ZoneInfo:
                    try:
                        value = value.astimezone(ZoneInfo(str(tz_name)))
                    except Exception:
                        pass
                else:
                    value = timezone.localtime(value)
            format_value = formatter.get("format")
            if format_value:
                return formats.date_format(value, format_value)
            return formats.date_format(
                value,
                "DATETIME_FORMAT" if isinstance(value, datetime) else "DATE_FORMAT",
            )
        if formatter_type == "number":
            if not isinstance(value, (int, float, Decimal)):
                return value
            decimal_pos = formatter.get("decimal_pos", None)
            use_l10n = bool(formatter.get("use_l10n", True))
            return formats.number_format(
                value, decimal_pos=decimal_pos, use_l10n=use_l10n
            )
        return value

    def parse_field_config(
        self, field_config: Union[str, dict[str, str]]
    ) -> dict[str, str]:
        """
        Parse field configuration to extract accessor and title.

        Args:
            field_config: Field configuration (string or dict)

        Returns:
            Dict with 'accessor' and 'title' keys
        """
        if isinstance(field_config, str):
            # String format: use field name as accessor and get verbose name as title
            accessor = field_config
            title = self.get_field_verbose_name(accessor)
            return {"accessor": accessor, "title": title}

        elif isinstance(field_config, dict):
            # Dict format: use provided accessor and title
            accessor = field_config.get("accessor", "")
            title = field_config.get("title", accessor)
            return {"accessor": accessor, "title": title}

        else:
            # Invalid format, use string representation
            accessor = str(field_config)
            title = accessor
            return {"accessor": accessor, "title": title}

    def get_field_verbose_name(self, field_path: str) -> str:
        """
        Get the verbose name for a field path, handling nested fields.

        Args:
            field_path: Field path (e.g., 'title', 'author.username')

        Returns:
            Verbose name of the field
        """
        try:
            parts = field_path.split(".")
            current_model = self.model
            verbose_name = field_path  # Default fallback

            for i, part in enumerate(parts):
                try:
                    field = current_model._meta.get_field(part)

                    if i == len(parts) - 1:  # Last part
                        verbose_name = getattr(field, "verbose_name", part)
                    else:
                        # Navigate to related model
                        if hasattr(field, "related_model"):
                            current_model = field.related_model
                        else:
                            break

                except Exception:
                    # Field not found, use the part name
                    verbose_name = part
                    break

            return str(verbose_name).title()

        except Exception as e:
            self.logger.debug(f"Could not get verbose name for {field_path}: {e}")
            return field_path.replace("_", " ").title()

    def get_field_headers(self, fields: list[Union[str, dict[str, str]]]) -> list[str]:
        """
        Generate column headers for the export with flexible field format support.

        Args:
            fields: List of field definitions (string or dict format)

        Returns:
            List of column headers
        """
        headers = []

        for field_config in fields:
            parsed_field = self.parse_field_config(field_config)
            headers.append(parsed_field["title"])

        return headers

    def _extract_field_data(self, obj, fields):
        """
        Extract field data from a model instance based on field configurations.

        Args:
            obj: Django model instance
            fields: List of field configurations (string or dict format)

        Returns:
            List of field values for the instance
        """
        row_data = []

        for field_config in fields:
            if isinstance(field_config, str):
                # String format: accessor is the field name
                accessor = field_config
            elif isinstance(field_config, dict):
                # Dict format: get accessor from dict
                accessor = field_config["accessor"]
            else:
                # Invalid format, use empty string
                row_data.append("")
                continue

            try:
                value = self.get_field_value(obj, accessor)
                row_data.append(value)
            except Exception as e:
                # Log error and use empty string as fallback
                logging.getLogger(__name__).warning(
                    f"Error extracting field '{accessor}': {e}"
                )
                row_data.append("")

        return row_data

    def _get_field_headers(self, fields):
        """
        Generate field headers from field configurations.

        Args:
            fields: List of field configurations (string or dict format)

        Returns:
            List of header strings for the export file
        """
        headers = []

        for field_config in fields:
            if isinstance(field_config, str):
                # String format: use verbose_name or field name as title
                accessor = field_config
                title = self._get_verbose_name_for_accessor(accessor)
                headers.append(title)
            elif isinstance(field_config, dict):
                # Dict format: use provided title or fallback to verbose_name
                accessor = field_config["accessor"]
                if "title" in field_config:
                    title = field_config["title"]
                else:
                    title = self._get_verbose_name_for_accessor(accessor)
                headers.append(title)

        return headers

    def _get_verbose_name_for_accessor(self, accessor):
        """
        Get the verbose name for a field accessor, handling nested fields.

        Args:
            accessor: Field accessor string (e.g., 'title', 'author.username')

        Returns:
            String representing the verbose name or field name
        """
        try:
            # Split accessor into parts for nested field access
            parts = accessor.split(".")
            current_model = self.model
            verbose_name = None

            for i, part in enumerate(parts):
                try:
                    field = current_model._meta.get_field(part)

                    if i == len(parts) - 1:
                        # Last part - get verbose name
                        verbose_name = getattr(field, "verbose_name", part)
                        if hasattr(field, "related_model") and field.related_model:
                            # For foreign key fields, might want to include related model info
                            verbose_name = str(verbose_name).title()
                    else:
                        # Intermediate part - move to related model
                        if hasattr(field, "related_model") and field.related_model:
                            current_model = field.related_model
                        else:
                            # Can't traverse further, use remaining parts as fallback
                            remaining_parts = ".".join(parts[i:])
                            verbose_name = remaining_parts.replace("_", " ").title()
                            break

                except FieldDoesNotExist:
                    # Field doesn't exist, might be a method or property
                    # Use the part name as fallback
                    if i == len(parts) - 1:
                        verbose_name = part.replace("_", " ").title()
                    else:
                        # Can't traverse further, use remaining parts as fallback
                        remaining_parts = ".".join(parts[i:])
                        verbose_name = remaining_parts.replace("_", " ").title()
                        break

            return verbose_name or accessor.replace("_", " ").title()

        except Exception:
            # Fallback: use accessor with underscores replaced by spaces
            return accessor.replace("_", " ").title()

    def export_to_csv(
        self,
        fields: list[Union[str, dict[str, str]]],
        variables: Optional[dict[str, Any]] = None,
        ordering: Optional[Union[str, list[str]]] = None,
        max_rows: Optional[int] = None,
        parsed_fields: Optional[list[dict[str, str]]] = None,
        output: Optional[io.StringIO] = None,
        progress_callback: Optional[Callable[[int], None]] = None,
        chunk_size: Optional[int] = None,
        *,
        presets: Optional[List[str]] = None,
        distinct_on: Optional[List[str]] = None,
    ) -> str:
        """
        Export model data to CSV format with flexible field format support.

        Args:
            fields: List of field definitions (string or dict format)
            variables: Filter variables (expects {"where": {...}})
            ordering: Ordering expression(s)
            max_rows: Optional max rows cap
            presets: Optional list of preset names to apply from GraphQLMeta.filter_presets
            distinct_on: Optional list of field names for DISTINCT ON (PostgreSQL only)

        Returns:
            CSV content as string
        """
        output = output or io.StringIO()
        writer = csv.writer(output)

        # Parse field configurations
        if parsed_fields is None:
            parsed_fields = self.validate_fields(
                fields, export_settings=self.export_settings
            )

        # Write headers
        headers = [parsed_field["title"] for parsed_field in parsed_fields]
        writer.writerow(headers)

        # Write data rows
        queryset = self.get_queryset(
            variables,
            ordering,
            fields=[field["accessor"] for field in parsed_fields],
            max_rows=max_rows,
            presets=presets,
            skip_validation=True,  # Already validated at view level
            distinct_on=distinct_on,
        )

        if chunk_size is None:
            chunk_size = int(self.export_settings.get("csv_chunk_size", 1000))
        if chunk_size <= 0:
            chunk_size = 1000

        processed = 0
        for instance in queryset.iterator(chunk_size=chunk_size):
            row = []
            for parsed_field in parsed_fields:
                accessor = parsed_field["accessor"]
                value = self.get_field_value(instance, accessor)
                row.append(value)
            writer.writerow(row)
            processed += 1
            if progress_callback and processed % chunk_size == 0:
                progress_callback(processed)

        return output.getvalue() if isinstance(output, io.StringIO) else ""

    def export_to_excel(
        self,
        fields: list[Union[str, dict[str, str]]],
        variables: Optional[dict[str, Any]] = None,
        ordering: Optional[Union[str, list[str]]] = None,
        max_rows: Optional[int] = None,
        parsed_fields: Optional[list[dict[str, str]]] = None,
        output: Optional[io.BytesIO] = None,
        progress_callback: Optional[Callable[[int], None]] = None,
        *,
        presets: Optional[List[str]] = None,
        distinct_on: Optional[List[str]] = None,
    ) -> bytes:
        """
        Export model data to Excel format with professional styling.

        Args:
            fields: List of field definitions (string or dict format)
            variables: Filter variables (expects {"where": {...}})
            ordering: Ordering expression(s)
            max_rows: Optional max rows cap
            presets: Optional list of preset names to apply from GraphQLMeta.filter_presets
            distinct_on: Optional list of field names for DISTINCT ON (PostgreSQL only)

        Returns:
            Excel file content as bytes

        Raises:
            ExportError: If openpyxl is not available
        """
        if not EXCEL_AVAILABLE:
            raise ExportError(
                "Excel export requires openpyxl package. Install with: pip install openpyxl"
            )

        # Use non-write-only mode for better styling support
        write_only = bool(self.export_settings.get("excel_write_only", False))
        workbook = openpyxl.Workbook(write_only=write_only)
        worksheet = workbook.active if not write_only else workbook.create_sheet()
        worksheet.title = f"{self.model_name} Export"

        # Hide gridlines (only works in non-write-only mode)
        if not write_only:
            worksheet.sheet_view.showGridLines = False

        # Professional style definitions - headers have fill only, no borders
        header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        header_fill = PatternFill(
            start_color="2F5496", end_color="2F5496", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        # Data row styles (no borders)
        data_font = Font(size=10, name="Calibri")
        data_alignment = Alignment(vertical="center", wrap_text=False)

        # Row number column style
        row_num_font = Font(size=10, name="Calibri", color="666666")
        row_num_alignment = Alignment(horizontal="center", vertical="center")

        # Alternating row colors
        even_row_fill = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        )
        odd_row_fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )

        # Parse field configurations
        if parsed_fields is None:
            parsed_fields = self.validate_fields(
                fields, export_settings=self.export_settings
            )

        # Write headers - first column is "#" for row numbers
        headers = ["#"] + [parsed_field["title"] for parsed_field in parsed_fields]
        if write_only:
            header_row = []
            for header in headers:
                cell = WriteOnlyCell(worksheet, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                # No border on headers
                header_row.append(cell)
            worksheet.append(header_row)
        else:
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                # No border on headers

        # Write data rows
        queryset = self.get_queryset(
            variables,
            ordering,
            fields=[field["accessor"] for field in parsed_fields],
            max_rows=max_rows,
            presets=presets,
            skip_validation=True,  # Already validated at view level
            distinct_on=distinct_on,
        )

        processed = 0
        progress_every = int(
            (self.export_settings.get("async_jobs") or {}).get(
                "progress_update_rows", 500
            )
        )
        if progress_every <= 0:
            progress_every = 500

        # Track max width per column for auto-sizing (including # column)
        column_widths = [3] + [len(str(h)) for h in headers[1:]]  # # column starts narrow

        row_counter = 0  # Sequential row number for # column
        for row_num, instance in enumerate(queryset.iterator(), 2):
            row_counter += 1
            is_even_row = (row_num % 2) == 0
            row_fill = even_row_fill if is_even_row else odd_row_fill

            if write_only:
                row = []
                # First cell: row number
                num_cell = WriteOnlyCell(worksheet, value=row_counter)
                num_cell.font = row_num_font
                num_cell.fill = row_fill
                num_cell.alignment = row_num_alignment
                row.append(num_cell)
                # Data cells
                for col_idx, parsed_field in enumerate(parsed_fields):
                    accessor = parsed_field["accessor"]
                    value = self.get_field_value(instance, accessor)
                    cell = WriteOnlyCell(worksheet, value=value)
                    cell.font = data_font
                    cell.fill = row_fill
                    cell.alignment = data_alignment
                    row.append(cell)
                    # Track width (offset by 1 for # column)
                    val_len = len(str(value)) if value else 0
                    if val_len > column_widths[col_idx + 1]:
                        column_widths[col_idx + 1] = min(val_len, 50)
                worksheet.append(row)
            else:
                # First cell: row number
                num_cell = worksheet.cell(row=row_num, column=1, value=row_counter)
                num_cell.font = row_num_font
                num_cell.fill = row_fill
                num_cell.alignment = row_num_alignment
                # Data cells (start at column 2)
                for col_num, parsed_field in enumerate(parsed_fields, 2):
                    accessor = parsed_field["accessor"]
                    value = self.get_field_value(instance, accessor)
                    cell = worksheet.cell(row=row_num, column=col_num, value=value)
                    cell.font = data_font
                    cell.fill = row_fill
                    cell.alignment = data_alignment
                    # Track width (col_num - 1 maps to index in column_widths)
                    val_len = len(str(value)) if value else 0
                    if val_len > column_widths[col_num - 1]:
                        column_widths[col_num - 1] = min(val_len, 50)

            processed += 1
            if progress_callback and processed % progress_every == 0:
                progress_callback(processed)

        # Apply column widths and additional formatting (non write-only only)
        if not write_only:
            # Set column widths with padding
            for col_idx, width in enumerate(column_widths, 1):
                column_letter = get_column_letter(col_idx)
                if col_idx == 1:  # # column - fixed narrow width
                    worksheet.column_dimensions[column_letter].width = 6
                else:
                    worksheet.column_dimensions[column_letter].width = min(width + 3, 50)

            # Set row height for header
            worksheet.row_dimensions[1].height = 25

        # Save to bytes
        output = output or io.BytesIO()
        workbook.save(output)
        return output.getvalue() if isinstance(output, io.BytesIO) else b""


@method_decorator(csrf_exempt, name="dispatch")
@method_decorator(_jwt_required, name="dispatch")
class ExportView(View):
    """
    Django view for handling model export requests with JWT authentication.

    Accepts POST requests with JSON payload containing export parameters
    and returns downloadable Excel or CSV files. All requests must include
    a valid JWT token in the Authorization header.

    Authentication:
        Requires JWT token: Authorization: Bearer <token>

    Field Format Examples:
    - String: "title" (uses field name as accessor and verbose_name as title)
    - Dict: {"accessor": "author.username", "title": "Author Name"}

    Filter Examples:
    - Basic: {"status": "active", "is_published": true}
    - Quick search: {"quick": "search term"}
    - Date filters: {"created_date_today": true, "updated_date_this_week": true}
    - Custom filters: {"has_tags": true, "content_length": "medium"}
    """

    def post(self, request):
        """
        Handle POST request for model export (JWT protected).

        Expected JSON payload:
        {
            "app_name": "blog",
            "model_name": "Post",
            "file_extension": "xlsx",  // or "csv"
            "filename": "posts_export",  // optional
            "fields": [
                "title",
                "author.username",
                {"accessor": "slug", "title": "MySlug"}
            ],
            "ordering": ["-created_at"],  // optional list
            "variables": {  // optional GraphQL filter parameters
                "status": "active",
                "quick": "search term",
                "published_date_today": true
            }
        }

        Returns:
            HttpResponse with file download or JsonResponse with error
        """
        # Log authenticated user for audit purposes
        if hasattr(request, "user") and request.user.is_authenticated:
            logger.info(
                f"Export request from user: {request.user.username} (ID: {request.user.id})"
            )

        audit_details = {"action": "export"}

        try:
            export_settings = _get_export_settings()
            rate_limit_response = self._check_rate_limit(request, export_settings)
            if rate_limit_response is not None:
                self._log_export_event(
                    request,
                    success=False,
                    error_message="Rate limit exceeded",
                    details=audit_details,
                )
                return rate_limit_response

            # Parse JSON payload
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                self._log_export_event(
                    request,
                    success=False,
                    error_message="Invalid JSON payload",
                    details=audit_details,
                )
                return JsonResponse({"error": "Invalid JSON payload"}, status=400)

            if not isinstance(data, dict):
                self._log_export_event(
                    request,
                    success=False,
                    error_message="Payload must be an object",
                    details=audit_details,
                )
                return JsonResponse({"error": "Payload must be an object"}, status=400)

            template_name = data.get("template")
            if template_name:
                resolved = self._apply_export_template(
                    request, str(template_name), data, export_settings
                )
                if isinstance(resolved, JsonResponse):
                    self._log_export_event(
                        request,
                        success=False,
                        error_message="Template not permitted",
                        details=audit_details,
                    )
                    return resolved
                data = resolved
                audit_details["template"] = str(template_name)

            # Validate required parameters
            required_fields = ["app_name", "model_name", "file_extension", "fields"]
            for field in required_fields:
                if field not in data:
                    self._log_export_event(
                        request,
                        success=False,
                        error_message=f"Missing required field: {field}",
                        details=audit_details,
                    )
                    return JsonResponse(
                        {"error": f"Missing required field: {field}"}, status=400
                    )

            app_name = data["app_name"]
            model_name = data["model_name"]
            file_extension = str(data["file_extension"]).lower()
            fields = data["fields"]
            audit_details.update(
                {
                    "app_name": app_name,
                    "model_name": model_name,
                    "file_extension": file_extension,
                }
            )

            # Validate file extension
            if file_extension in ["excel", "xlsx"]:
                file_extension = "xlsx"
            if file_extension not in ["xlsx", "csv"]:
                self._log_export_event(
                    request,
                    success=False,
                    error_message='file_extension must be "xlsx" or "csv"',
                    details=audit_details,
                )
                return JsonResponse(
                    {"error": 'file_extension must be "xlsx" or "csv"'}, status=400
                )
            audit_details["file_extension"] = file_extension

            # Validate fields format
            if not isinstance(fields, list) or not fields:
                self._log_export_event(
                    request,
                    success=False,
                    error_message="fields must be a non-empty list",
                    details=audit_details,
                )
                return JsonResponse(
                    {"error": "fields must be a non-empty list"}, status=400
                )
            audit_details["fields_count"] = len(fields)

            # Validate field configurations
            for i, field_config in enumerate(fields):
                if isinstance(field_config, str):
                    continue  # String format is valid
                elif isinstance(field_config, dict):
                    if "accessor" not in field_config:
                        self._log_export_event(
                            request,
                            success=False,
                            error_message="Invalid field configuration",
                            details=audit_details,
                        )
                        return JsonResponse(
                            {
                                "error": f"Invalid field configuration at index {i}: dict format must contain 'accessor' key"
                            },
                            status=400,
                        )
                else:
                    self._log_export_event(
                        request,
                        success=False,
                        error_message="Invalid field configuration",
                        details=audit_details,
                    )
                    return JsonResponse(
                        {
                            "error": f"Invalid field configuration at index {i}: field must be string or dict with accessor/title"
                        },
                        status=400,
                    )

            # Optional parameters
            filename = data.get("filename")
            ordering = data.get("ordering")
            variables = data.get("variables") or {}
            presets = data.get("presets")  # List of preset names for filtering
            schema_name = data.get(
                "schema_name"
            )  # Schema name for multi-schema support
            distinct_on = data.get("distinct_on")  # List of fields for DISTINCT ON
            async_value = data.get("async", False)
            if async_value is not None and not isinstance(async_value, bool):
                self._log_export_event(
                    request,
                    success=False,
                    error_message="async must be a boolean",
                    details=audit_details,
                )
                return JsonResponse({"error": "async must be a boolean"}, status=400)
            async_request = bool(async_value)

            if not isinstance(variables, dict):
                self._log_export_event(
                    request,
                    success=False,
                    error_message="variables must be an object of filter parameters",
                    details=audit_details,
                )
                return JsonResponse(
                    {"error": "variables must be an object of filter parameters"},
                    status=400,
                )

            if ordering is not None and not isinstance(ordering, (list, tuple, str)):
                self._log_export_event(
                    request,
                    success=False,
                    error_message="ordering must be a string or list",
                    details=audit_details,
                )
                return JsonResponse(
                    {"error": "ordering must be a string or list"}, status=400
                )

            if presets is not None and not isinstance(presets, list):
                self._log_export_event(
                    request,
                    success=False,
                    error_message="presets must be a list of preset names",
                    details=audit_details,
                )
                return JsonResponse(
                    {"error": "presets must be a list of preset names"}, status=400
                )

            if schema_name is not None and not isinstance(schema_name, str):
                self._log_export_event(
                    request,
                    success=False,
                    error_message="schema_name must be a string",
                    details=audit_details,
                )
                return JsonResponse(
                    {"error": "schema_name must be a string"}, status=400
                )

            if distinct_on is not None and not isinstance(distinct_on, list):
                self._log_export_event(
                    request,
                    success=False,
                    error_message="distinct_on must be a list of field names",
                    details=audit_details,
                )
                return JsonResponse(
                    {"error": "distinct_on must be a list of field names"}, status=400
                )

            max_rows, max_rows_error = self._resolve_max_rows(data, export_settings)
            if max_rows_error is not None:
                self._log_export_event(
                    request,
                    success=False,
                    error_message="Invalid max_rows",
                    details=audit_details,
                )
                return max_rows_error
            audit_details["max_rows"] = max_rows

            # Generate default filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{model_name}_{timestamp}"
            filename = _sanitize_filename(str(filename))

            # Create exporter and generate file
            exporter = ModelExporter(
                app_name,
                model_name,
                export_settings=export_settings,
                schema_name=schema_name,
            )
            permission_response = self._enforce_model_permissions(
                request, exporter.model, export_settings
            )
            if permission_response is not None:
                audit_details["model_label"] = exporter.model._meta.label
                self._log_export_event(
                    request,
                    success=False,
                    error_message="Export not permitted",
                    details=audit_details,
                )
                return permission_response

            parsed_fields = exporter.validate_fields(
                fields,
                user=getattr(request, "user", None),
                export_settings=export_settings,
            )
            # Use the new validate_filter_input method with standardized "where" key
            where_input = variables.get("where", variables) if variables else None
            exporter.validate_filter_input(where_input, export_settings=export_settings)
            ordering_fields = exporter._normalize_ordering(ordering)
            ordering_value = ordering_fields or None

            async_settings = export_settings.get("async_jobs") or {}
            if async_request:
                if not async_settings.get("enable", False):
                    self._log_export_event(
                        request,
                        success=False,
                        error_message="Async export is disabled",
                        details=audit_details,
                    )
                    return JsonResponse(
                        {"error": "Async export is disabled"}, status=400
                    )

                job_response = self._enqueue_export_job(
                    request=request,
                    exporter=exporter,
                    parsed_fields=parsed_fields,
                    variables=variables,
                    ordering=ordering_value,
                    max_rows=max_rows,
                    filename=filename,
                    file_extension=file_extension,
                    export_settings=export_settings,
                )
                audit_details["async_job"] = True
                self._log_export_event(
                    request,
                    success=True,
                    details=audit_details,
                )
                return job_response

            if file_extension == "xlsx":
                content = exporter.export_to_excel(
                    fields,
                    variables,
                    ordering_value,
                    max_rows=max_rows,
                    parsed_fields=parsed_fields,
                    presets=presets,
                    distinct_on=distinct_on,
                )
                content_type = (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                file_ext = "xlsx"
            else:  # csv
                if export_settings.get(
                    "enforce_streaming_csv", True
                ) or export_settings.get("stream_csv", True):
                    audit_details["stream_csv"] = True
                    self._log_export_event(
                        request,
                        success=True,
                        details=audit_details,
                    )
                    return self._stream_csv_response(
                        exporter=exporter,
                        parsed_fields=parsed_fields,
                        variables=variables,
                        ordering=ordering_value,
                        max_rows=max_rows,
                        filename=filename,
                        chunk_size=int(export_settings.get("csv_chunk_size", 1000)),
                        presets=presets,
                        distinct_on=distinct_on,
                    )
                content = exporter.export_to_csv(
                    fields,
                    variables,
                    ordering_value,
                    max_rows=max_rows,
                    parsed_fields=parsed_fields,
                    presets=presets,
                    distinct_on=distinct_on,
                )
                content_type = "text/csv; charset=utf-8"
                file_ext = "csv"
                audit_details["stream_csv"] = False

            # Create HTTP response with file download
            response = HttpResponse(content, content_type=content_type)
            response["Content-Disposition"] = (
                f'attachment; filename="{filename}.{file_ext}"'
            )
            response["Content-Length"] = len(content)

            logger.info(
                f"Successfully exported {model_name} data to {file_extension} format"
            )
            self._log_export_event(
                request,
                success=True,
                details=audit_details,
            )
            return response

        except ExportError as e:
            logger.error(f"Export error: {e}")
            message = str(e)
            status = 403 if "denied" in message.lower() else 400
            self._log_export_event(
                request,
                success=False,
                error_message=message,
                details=audit_details,
            )
            return JsonResponse({"error": message}, status=status)
        except Exception as e:
            logger.error(f"Unexpected error during export: {e}")
            self._log_export_event(
                request,
                success=False,
                error_message="Internal server error",
                details=audit_details,
            )
            return JsonResponse({"error": "Internal server error"}, status=500)

    def _log_export_event(
        self,
        request: Any,
        *,
        success: bool,
        error_message: Optional[str] = None,
        details: Optional[dict[str, Any]] = None,
    ) -> None:
        """Log an audit event for export activity."""
        if not log_audit_event or not AuditEventType:
            return

        audit_details = {"action": "export"}
        if details:
            audit_details.update(details)

        log_audit_event(
            request,
            AuditEventType.DATA_ACCESS,
            success=success,
            error_message=error_message,
            additional_data=audit_details,
        )

    def _get_rate_limit_identifier(
        self, request: Any, export_settings: dict[str, Any]
    ) -> str:
        """Resolve the rate limit identifier for the request."""
        user = getattr(request, "user", None)
        if user and getattr(user, "is_authenticated", False):
            return f"user:{user.id}"

        rate_limit = export_settings.get("rate_limit") or {}
        trusted_proxies = rate_limit.get("trusted_proxies") or []
        remote_addr = request.META.get("REMOTE_ADDR", "")
        ip_address = remote_addr or "unknown"

        if self._is_trusted_proxy(remote_addr, trusted_proxies):
            forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR", "")
            if forwarded_for:
                ip_address = forwarded_for.split(",")[0].strip()

        return f"ip:{ip_address}"

    def _is_trusted_proxy(
        self, remote_addr: str, trusted_proxies: Iterable[str]
    ) -> bool:
        """Check if the remote address is in the trusted proxy list."""
        if not remote_addr:
            return False
        for proxy in trusted_proxies:
            proxy = str(proxy).strip()
            if not proxy:
                continue
            if "/" in proxy:
                try:
                    if ipaddress.ip_address(remote_addr) in ipaddress.ip_network(
                        proxy, strict=False
                    ):
                        return True
                except ValueError:
                    continue
            if remote_addr == proxy:
                return True
        return False

    def _check_rate_limit(
        self, request: Any, export_settings: dict[str, Any]
    ) -> Optional[JsonResponse]:
        """Apply a basic rate limit using Django cache."""
        config = export_settings.get("rate_limit") or {}
        if not config.get("enable", True):
            return None

        window_seconds = int(config.get("window_seconds", 60))
        max_requests = int(config.get("max_requests", 30))
        identifier = self._get_rate_limit_identifier(request, export_settings)
        cache_key = f"rail:export_rl:{identifier}"

        count = cache.get(cache_key)
        if count is None:
            cache.add(cache_key, 1, timeout=window_seconds)
            return None

        if int(count) >= max_requests:
            return JsonResponse(
                {"error": "Rate limit exceeded", "retry_after": window_seconds},
                status=429,
            )

        try:
            cache.incr(cache_key)
        except ValueError:
            cache.set(cache_key, int(count) + 1, timeout=window_seconds)
        return None

    def _resolve_max_rows(
        self, data: dict[str, Any], export_settings: dict[str, Any]
    ) -> tuple[Optional[int], Optional[JsonResponse]]:
        """Resolve max rows with request override and config cap."""
        config_max_rows = export_settings.get("max_rows", None)
        requested_max = data.get("max_rows", data.get("limit"))

        if config_max_rows is not None:
            try:
                config_max_rows = int(config_max_rows)
            except (TypeError, ValueError):
                config_max_rows = None
        if config_max_rows is not None and config_max_rows <= 0:
            config_max_rows = None

        if requested_max is None:
            return config_max_rows, None

        try:
            requested_max = int(requested_max)
        except (TypeError, ValueError):
            return None, JsonResponse(
                {"error": "max_rows must be an integer"}, status=400
            )

        if requested_max <= 0:
            return config_max_rows, None

        if config_max_rows is None:
            return requested_max, None

        return min(requested_max, config_max_rows), None

    def _apply_export_template(
        self,
        request: Any,
        template_name: str,
        data: dict[str, Any],
        export_settings: dict[str, Any],
    ) -> Union[dict[str, Any], JsonResponse]:
        """Merge a named export template into the request payload."""
        templates = _get_export_templates(export_settings)
        template = templates.get(template_name)
        if not template:
            return JsonResponse(
                {"error": "Export template not found", "template": template_name},
                status=404,
            )
        if not isinstance(template, dict):
            return JsonResponse(
                {"error": "Export template is invalid", "template": template_name},
                status=400,
            )
        if not self._template_allowed(request, template):
            return JsonResponse(
                {"error": "Export template not permitted", "template": template_name},
                status=403,
            )

        merged = dict(template)
        merged["template"] = template_name

        allow_overrides = template.get(
            "allow_overrides", ["variables", "filename", "max_rows", "ordering"]
        )
        if isinstance(allow_overrides, (list, tuple, set)):
            for key in allow_overrides:
                if key in data:
                    if key == "variables" and isinstance(data[key], dict):
                        base_vars = dict(template.get("variables") or {})
                        base_vars.update(data[key])
                        merged[key] = base_vars
                    else:
                        merged[key] = data[key]

        return merged

    def _template_allowed(self, request: Any, template: dict[str, Any]) -> bool:
        """Check whether a user can access a template."""
        user = getattr(request, "user", None)
        if not user or not getattr(user, "is_authenticated", False):
            return False
        if getattr(user, "is_superuser", False):
            return True

        if template.get("shared", False):
            return True

        required_permissions = template.get("required_permissions") or template.get(
            "permissions"
        )
        if required_permissions:
            if isinstance(required_permissions, str):
                required_permissions = [required_permissions]
            if any(user.has_perm(perm) for perm in required_permissions):
                return True
            return False

        allowed_groups = template.get("allowed_groups") or []
        if allowed_groups:
            if isinstance(allowed_groups, str):
                allowed_groups = [allowed_groups]
            if user.groups.filter(name__in=list(allowed_groups)).exists():
                return True
            return False

        allowed_users = template.get("allowed_users") or []
        if allowed_users:
            if isinstance(allowed_users, (str, int)):
                allowed_users = [allowed_users]
            if str(user.id) in {str(value) for value in allowed_users}:
                return True
            if getattr(user, "username", None) in {
                str(value) for value in allowed_users
            }:
                return True
            return False

        return False

    def _enqueue_export_job(
        self,
        *,
        request: Any,
        exporter: ModelExporter,
        parsed_fields: list[dict[str, str]],
        variables: dict[str, Any],
        ordering: Optional[Union[str, list[str]]],
        max_rows: Optional[int],
        filename: str,
        file_extension: str,
        export_settings: dict[str, Any],
    ) -> JsonResponse:
        """Create and enqueue an export job for async processing."""
        async_settings = export_settings.get("async_jobs") or {}
        backend = str(async_settings.get("backend", "thread")).lower()
        expires_seconds = int(async_settings.get("expires_seconds", 3600))

        job_id = str(uuid.uuid4())
        now = timezone.now()
        job = {
            "id": job_id,
            "status": "pending",
            "created_at": now.isoformat(),
            "expires_at": (now + timedelta(seconds=expires_seconds)).isoformat(),
            "owner_id": getattr(getattr(request, "user", None), "id", None),
            "file_extension": file_extension,
            "filename": filename,
            "processed_rows": 0,
            "total_rows": None,
        }

        payload = {
            "app_name": exporter.app_name,
            "model_name": exporter.model_name,
            "file_extension": file_extension,
            "filename": filename,
            "fields": parsed_fields,
            "parsed_fields": parsed_fields,
            "variables": variables,
            "ordering": ordering,
            "max_rows": max_rows,
        }

        _set_export_job(job_id, job, timeout=expires_seconds)
        cache.set(_export_job_payload_key(job_id), payload, timeout=expires_seconds)

        if backend == "thread":
            thread = threading.Thread(
                target=_run_export_job, args=(job_id,), daemon=True
            )
            thread.start()
        elif backend == "celery":
            if not export_job_task:
                _update_export_job(
                    job_id,
                    {"status": "failed", "error": "Celery is not available"},
                    timeout=expires_seconds,
                )
                return JsonResponse(
                    {"error": "Celery backend not available"}, status=500
                )
            export_job_task.delay(job_id)
        elif backend == "rq":
            try:
                import django_rq
            except Exception:
                _update_export_job(
                    job_id,
                    {"status": "failed", "error": "RQ is not available"},
                    timeout=expires_seconds,
                )
                return JsonResponse({"error": "RQ backend not available"}, status=500)
            queue_name = async_settings.get("queue", "default")
            queue = django_rq.get_queue(queue_name)
            queue.enqueue(_run_export_job, job_id)
        else:
            _update_export_job(
                job_id,
                {"status": "failed", "error": "Unknown async backend"},
                timeout=expires_seconds,
            )
            return JsonResponse({"error": "Unknown async backend"}, status=500)

        base_path = request.path.rstrip("/")
        status_path = f"{base_path}/jobs/{job_id}/"
        download_path = f"{base_path}/jobs/{job_id}/download/"
        return JsonResponse(
            {
                "job_id": job_id,
                "status": "pending",
                "status_url": request.build_absolute_uri(status_path),
                "download_url": request.build_absolute_uri(download_path),
                "expires_in": expires_seconds,
            },
            status=202,
        )

    def _enforce_model_permissions(
        self, request: Any, model: type, export_settings: dict[str, Any]
    ) -> Optional[JsonResponse]:
        """Check model allowlist and permissions."""
        if not _is_model_allowed(model, export_settings):
            return JsonResponse(
                {"error": "Model export not allowed", "model": model._meta.label},
                status=403,
            )

        user = getattr(request, "user", None)
        if not user or not getattr(user, "is_authenticated", False):
            return JsonResponse(
                {"error": "Authentication required for export"}, status=401
            )

        if not export_settings.get("require_model_permissions", True):
            return None

        if getattr(user, "is_superuser", False):
            return None

        required_permissions = export_settings.get("required_permissions") or []
        if required_permissions:
            if not any(user.has_perm(perm) for perm in required_permissions):
                return JsonResponse(
                    {"error": "Insufficient permissions for export"}, status=403
                )
            return None

        if permission_manager and OperationType:
            result = permission_manager.check_operation_permission(
                user, model._meta.label_lower, OperationType.READ
            )
            if not result.allowed:
                return JsonResponse(
                    {
                        "error": "Insufficient permissions for export",
                        "detail": result.reason,
                    },
                    status=403,
                )

        view_perm = f"{model._meta.app_label}.view_{model._meta.model_name}"
        if not user.has_perm(view_perm):
            return JsonResponse(
                {"error": "Insufficient permissions for export"}, status=403
            )

        return None

    def _stream_csv_response(
        self,
        *,
        exporter: ModelExporter,
        parsed_fields: list[dict[str, str]],
        variables: dict[str, Any],
        ordering: Optional[Union[str, list[str]]],
        max_rows: Optional[int],
        filename: str,
        chunk_size: int,
        presets: Optional[List[str]] = None,
        distinct_on: Optional[List[str]] = None,
    ) -> StreamingHttpResponse:
        """Stream a CSV export response."""
        headers = [field["title"] for field in parsed_fields]
        accessors = [field["accessor"] for field in parsed_fields]

        if chunk_size <= 0:
            chunk_size = 1000

        queryset = exporter.get_queryset(
            variables,
            ordering,
            fields=accessors,
            max_rows=max_rows,
            presets=presets,
            skip_validation=True,  # Already validated at view level
            distinct_on=distinct_on,
        ).iterator(chunk_size=chunk_size)

        def row_generator():
            output = io.StringIO()
            writer = csv.writer(output)

            writer.writerow(headers)
            yield output.getvalue()
            output.seek(0)
            output.truncate(0)

            for instance in queryset:
                row = [
                    exporter.get_field_value(instance, accessor)
                    for accessor in accessors
                ]
                writer.writerow(row)
                yield output.getvalue()
                output.seek(0)
                output.truncate(0)

        response = StreamingHttpResponse(
            row_generator(), content_type="text/csv; charset=utf-8"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
        return response

    def get(self, request):
        """
        Handle GET request - return API documentation (JWT protected).
        """
        # Log authenticated user for audit purposes
        if hasattr(request, "user") and request.user.is_authenticated:
            logger.info(
                f"Export API documentation request from user: {request.user.username}"
            )

        documentation = {
            "endpoint": "/export",
            "method": "POST",
            "authentication": "JWT token required in Authorization header",
            "description": "Export Django model data to Excel or CSV format with GraphQL filter integration",
            "required_headers": {
                "Authorization": "Bearer <jwt_token>",
                "Content-Type": "application/json",
            },
            "required_parameters": {
                "app_name": "string - Name of the Django app containing the model",
                "model_name": "string - Name of the Django model to export",
                "file_extension": 'string - Either "xlsx" or "csv"',
                "fields": "array - List of field configurations (string or dict format)",
            },
            "optional_parameters": {
                "filename": "string - Custom filename (default: ModelName_timestamp)",
                "ordering": "array - List of Django ORM ordering expressions",
                "variables": "object - GraphQL filter parameters",
                "max_rows": "integer - Optional cap for rows exported (bounded by server settings)",
                "template": "string - Named export template configured on the server",
                "async": "boolean - Run export asynchronously and return a job id",
            },
            "field_formats": {
                "string": "Uses field name as accessor and verbose_name as title",
                "dict": "Must contain 'accessor' key, optionally 'title' key",
            },
            "constraints": {
                "accessors": "Must be allowlisted and use dot notation",
                "filters": "Must be allowlisted and respect filter guardrails",
                "ordering": "Must be allowlisted",
            },
            "filter_examples": {
                "basic": {"status": "active", "is_published": True},
                "quick_search": {"quick": "search term"},
                "date_filters": {
                    "created_date_today": True,
                    "updated_date_this_week": True,
                },
                "custom_filters": {"has_tags": True, "content_length": "medium"},
            },
            "example_request": {
                "url": "/api/v1/export/",
                "method": "POST",
                "headers": {
                    "Authorization": "Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9...",
                    "Content-Type": "application/json",
                },
                "payload": {
                    "app_name": "blog",
                    "model_name": "Post",
                    "file_extension": "xlsx",
                    "filename": "blog_posts_export",
                    "fields": [
                        "title",
                        "author.username",
                        {"accessor": "slug", "title": "MySlug"},
                    ],
                    "ordering": ["-created_at"],
                    "variables": {
                        "status": "active",
                        "quick": "search term",
                        "published_date_today": True,
                    },
                },
            },
            "async_example": {
                "payload": {
                    "template": "recent_posts",
                    "async": True,
                    "variables": {"status": "published"},
                }
            },
            "authentication_errors": {
                "401": "Missing or invalid JWT token",
                "403": "Token valid but insufficient permissions",
            },
            "async_endpoints": {
                "status": "GET /api/v1/export/jobs/<job_id>/",
                "download": "GET /api/v1/export/jobs/<job_id>/download/",
            },
        }

        return JsonResponse(documentation, json_dumps_params={"indent": 2})


@method_decorator(_jwt_required, name="dispatch")
class ExportJobStatusView(View):
    """Return export job status details."""

    def get(self, request, job_id):
        job_id = str(job_id)
        job = _get_export_job(job_id)
        if not job:
            raise Http404("Export job not found")
        if not _job_access_allowed(request, job):
            return JsonResponse({"error": "Export job not permitted"}, status=403)

        expires_at = _parse_iso_datetime(job.get("expires_at"))
        if expires_at and timezone.now() > expires_at:
            _cleanup_export_job_files(job)
            _delete_export_job(job_id)
            return JsonResponse({"error": "Export job expired"}, status=410)

        base_path = request.path.rstrip("/")
        download_path = f"{base_path}/download/"
        response = {
            "job_id": job_id,
            "status": job.get("status"),
            "processed_rows": job.get("processed_rows"),
            "total_rows": job.get("total_rows"),
            "error": job.get("error"),
            "created_at": job.get("created_at"),
            "completed_at": job.get("completed_at"),
            "expires_at": job.get("expires_at"),
        }
        if job.get("status") == "completed":
            response["download_url"] = request.build_absolute_uri(download_path)
        return JsonResponse(response)


@method_decorator(_jwt_required, name="dispatch")
class ExportJobDownloadView(View):
    """Download completed export job files."""

    def get(self, request, job_id):
        job_id = str(job_id)
        job = _get_export_job(job_id)
        if not job:
            raise Http404("Export job not found")
        if not _job_access_allowed(request, job):
            return JsonResponse({"error": "Export job not permitted"}, status=403)

        expires_at = _parse_iso_datetime(job.get("expires_at"))
        if expires_at and timezone.now() > expires_at:
            _cleanup_export_job_files(job)
            _delete_export_job(job_id)
            return JsonResponse({"error": "Export job expired"}, status=410)

        if job.get("status") != "completed":
            return JsonResponse({"error": "Export job not completed"}, status=409)

        file_path = job.get("file_path")
        if not file_path or not Path(file_path).exists():
            raise Http404("Export file not found")

        filename = _sanitize_filename(str(job.get("filename") or "export"))
        extension = job.get("file_extension") or "csv"
        response = FileResponse(
            open(file_path, "rb"),
            content_type=job.get("content_type", "application/octet-stream"),
            as_attachment=True,
            filename=f"{filename}.{extension}",
        )
        return response


# URL configuration helper
def get_export_urls():
    """
    Helper function to get URL patterns for the export functionality.

    Usage in urls.py:
        from rail_django.extensions.exporting import get_export_urls

        urlpatterns = [
            # ... other patterns
        ] + get_export_urls()

    Returns:
        List of URL patterns (export + async job endpoints)
    """
    from django.urls import path

    if not JWT_REQUIRED_AVAILABLE:
        raise ImproperlyConfigured(
            "Export endpoints require JWT auth; auth_decorators is missing."
        )

    return [
        path("export/", ExportView.as_view(), name="model_export"),
        path(
            "export/jobs/<uuid:job_id>/",
            ExportJobStatusView.as_view(),
            name="export_job_status",
        ),
        path(
            "export/jobs/<uuid:job_id>/download/",
            ExportJobDownloadView.as_view(),
            name="export_job_download",
        ),
    ]


# Utility functions for programmatic use
def export_model_to_csv(
    app_name: str,
    model_name: str,
    fields: list[Union[str, dict[str, str]]],
    variables: Optional[dict[str, Any]] = None,
    ordering: Optional[Union[str, list[str]]] = None,
    *,
    export_settings: Optional[dict[str, Any]] = None,
) -> str:
    """
    Programmatically export model data to CSV format with flexible field format support.

    Args:
        app_name: Name of the Django app
        model_name: Name of the model
        fields: List of field definitions (string or dict format)
        variables: Filter variables
        ordering: Ordering expression
        export_settings: Optional export configuration override

    Returns:
        CSV content as string
    """
    exporter = ModelExporter(app_name, model_name, export_settings=export_settings)
    return exporter.export_to_csv(fields, variables, ordering)


def export_model_to_excel(
    app_name: str,
    model_name: str,
    fields: list[Union[str, dict[str, str]]],
    variables: Optional[dict[str, Any]] = None,
    ordering: Optional[Union[str, list[str]]] = None,
    *,
    export_settings: Optional[dict[str, Any]] = None,
) -> bytes:
    """
    Programmatically export model data to Excel format with flexible field format support.

    Args:
        app_name: Name of the Django app
        model_name: Name of the model
        fields: List of field definitions (string or dict format)
        variables: Filter variables
        ordering: Ordering expression
        export_settings: Optional export configuration override

    Returns:
        Excel file content as bytes
    """
    exporter = ModelExporter(app_name, model_name, export_settings=export_settings)
    return exporter.export_to_excel(fields, variables, ordering)
