"""
Excel templating helpers built on top of openpyxl with pluggable styling.

This module lets models expose downloadable Excel files by decorating a model method with
`@model_excel_template`. The decorator registers a dynamic Django view that:
- Finds the related model instance (by PK passed in the URL)
- Calls the decorated method to get data (single sheet or multi-sheet)
- Applies optional style configuration (headers, cells, column widths, etc.)
- Streams the generated Excel file with the configured styling
- Supports async jobs, catalog endpoints, and optional rate limiting

Usage inside a model:
    from rail_django.extensions.excel_export import model_excel_template

    class Product(models.Model):
        name = models.CharField(max_length=100)
        price = models.DecimalField(max_digits=10, decimal_places=2)

        @model_excel_template(
            url="products/export",
            title="Product Export",
            config={
                "header_style": {"bold": True, "fill_color": "4472C4", "font_color": "FFFFFF"},
                "freeze_panes": True,
                "column_widths": "auto",
            }
        )
        def export_products(self):
            products = Product.objects.all()
            return [
                ["Name", "Price", "Created At"],  # Headers
                *[[p.name, p.price, p.created_at] for p in products]
            ]

The view is automatically available at:
    /api/excel/<template_path>/<pk>/

If `url` is omitted, the default path is: <app_label>/<model_name>/<function_name>.
Default style configuration comes from `settings.RAIL_DJANGO_GRAPHQL_EXCEL_EXPORT`.
"""

import hashlib
import inspect
import io
import ipaddress
import json
import logging
import re
import tempfile
import threading
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Tuple, Type, Union

from django.conf import settings
from django.contrib.auth import get_user_model
from django.contrib.auth.models import AnonymousUser
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.db import models
from django.db.models.signals import class_prepared
from django.http import FileResponse, Http404, HttpRequest, HttpResponse, JsonResponse
from django.urls import path
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.csrf import csrf_exempt

# Optional openpyxl support
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet

    OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover - optional dependency
    openpyxl = None
    Workbook = None
    WriteOnlyCell = None
    Alignment = None
    Border = None
    Font = None
    PatternFill = None
    Side = None
    get_column_letter = None
    Worksheet = None
    OPENPYXL_AVAILABLE = False

# Optional GraphQL metadata and permission helpers
try:
    from rail_django.core.meta import get_model_graphql_meta
except ImportError:  # pragma: no cover - fallback during early startup
    get_model_graphql_meta = None

try:
    from rail_django.extensions.auth import get_user_from_token
except ImportError:  # pragma: no cover - fallback when auth not available
    get_user_from_token = None

try:
    from rail_django.extensions.permissions import (
        OperationType,
        permission_manager,
    )
except ImportError:  # pragma: no cover - optional permission subsystem
    OperationType = None
    permission_manager = None

try:
    from rail_django.security.rbac import role_manager
except ImportError:  # pragma: no cover - security optional
    role_manager = None

# Optional JWT protection (mirrors export endpoints)
try:
    from .auth_decorators import jwt_required
except ImportError:
    jwt_required = None

try:
    from .audit import AuditEventType, log_audit_event
except ImportError:
    AuditEventType = None
    log_audit_event = None

logger = logging.getLogger(__name__)

# Type aliases for data formats
ExcelRowData = List[Any]
ExcelSheetData = List[ExcelRowData]
ExcelMultiSheetData = Dict[str, ExcelSheetData]
ExcelData = Union[ExcelSheetData, ExcelMultiSheetData]

EXCEL_RATE_LIMIT_DEFAULTS = {
    "enable": True,
    "window_seconds": 60,
    "max_requests": 30,
    "trusted_proxies": [],
}

EXCEL_CACHE_DEFAULTS = {
    "enable": False,
    "timeout_seconds": 300,
    "vary_on_user": True,
    "key_prefix": "rail:excel_cache",
}

EXCEL_ASYNC_DEFAULTS = {
    "enable": False,
    "backend": "thread",
    "expires_seconds": 3600,
    "storage_dir": None,
    "queue": "default",
    "track_progress": False,
    "webhook_url": None,
    "webhook_headers": {},
    "webhook_timeout_seconds": 10,
}

EXCEL_CATALOG_DEFAULTS = {
    "enable": True,
    "require_authentication": True,
    "filter_by_access": True,
    "include_config": False,
    "include_permissions": True,
}

DEFAULT_HEADER_STYLE = {
    "bold": True,
    "fill_color": "4472C4",
    "font_color": "FFFFFF",
    "font_size": 11,
    "alignment": "center",
}

DEFAULT_CELL_STYLE = {
    "font_size": 11,
    "alignment": "left",
    "wrap_text": False,
}

DEFAULT_ALTERNATING_ROW_STYLE = {
    "enable": True,
    "even_fill_color": "F2F2F2",
    "odd_fill_color": "FFFFFF",
}

DEFAULT_BORDER_STYLE = {
    "enable": True,
    "color": "D4D4D4",
    "style": "thin",
}


def _merge_dict(defaults: Dict[str, Any], overrides: Any) -> Dict[str, Any]:
    """Shallow-merge dict settings with safe fallbacks."""
    merged = dict(defaults)
    if isinstance(overrides, dict):
        merged.update(overrides)
    return merged


def _excel_export_settings() -> Dict[str, Any]:
    """
    Safely read the Excel export defaults from settings.

    Returns:
        A dictionary with style defaults and configuration.
    """
    return getattr(settings, "RAIL_DJANGO_GRAPHQL_EXCEL_EXPORT", {})


def _excel_dict(key: str, defaults: Dict[str, Any]) -> Dict[str, Any]:
    return _merge_dict(defaults, _excel_export_settings().get(key))


def _excel_rate_limit() -> Dict[str, Any]:
    return _excel_dict("rate_limit", EXCEL_RATE_LIMIT_DEFAULTS)


def _excel_cache() -> Dict[str, Any]:
    return _excel_dict("cache", EXCEL_CACHE_DEFAULTS)


def _excel_async() -> Dict[str, Any]:
    return _excel_dict("async_jobs", EXCEL_ASYNC_DEFAULTS)


def _excel_catalog() -> Dict[str, Any]:
    return _excel_dict("catalog", EXCEL_CATALOG_DEFAULTS)


def _excel_expose_errors() -> bool:
    return bool(_excel_export_settings().get("expose_errors", settings.DEBUG))


def _url_prefix() -> str:
    """Return URL prefix under /api/ where Excel templates are exposed."""
    return _excel_export_settings().get("url_prefix", "excel")


def _default_excel_config() -> Dict[str, Any]:
    """
    Provide default styling that can be overridden per template.

    Returns:
        Dict of Excel configuration values.
    """
    defaults = {
        "sheet_name": "Sheet1",
        "freeze_panes": True,
        "column_widths": "auto",
        "header_style": DEFAULT_HEADER_STYLE.copy(),
        "cell_style": DEFAULT_CELL_STYLE.copy(),
        "alternating_rows": DEFAULT_ALTERNATING_ROW_STYLE.copy(),
        "borders": DEFAULT_BORDER_STYLE.copy(),
        "number_formats": {},
        "date_format": "YYYY-MM-DD",
        "datetime_format": "YYYY-MM-DD HH:MM:SS",
        "decimal_format": "#,##0.00",
        "auto_filter": True,
    }
    settings_overrides = _excel_export_settings().get("default_config", {})
    return {**defaults, **settings_overrides}


@dataclass
class ExcelTemplateMeta:
    """Raw decorator metadata attached to a model method."""

    url_path: Optional[str]
    config: Dict[str, Any] = field(default_factory=dict)
    roles: Sequence[str] = field(default_factory=tuple)
    permissions: Sequence[str] = field(default_factory=tuple)
    guard: Optional[str] = None
    require_authentication: bool = True
    title: Optional[str] = None
    allow_client_data: bool = False
    client_data_fields: Sequence[str] = field(default_factory=tuple)


@dataclass
class ExcelTemplateDefinition:
    """Runtime representation of a registered Excel template."""

    model: Optional[Type[models.Model]]
    method_name: Optional[str]
    handler: Optional[Callable[..., Any]]
    source: str
    url_path: str
    config: Dict[str, Any]
    roles: Sequence[str]
    permissions: Sequence[str]
    guard: Optional[str]
    require_authentication: bool
    title: str
    allow_client_data: bool
    client_data_fields: Sequence[str]


@dataclass
class ExcelTemplateAccessDecision:
    """Represents whether a user can access an Excel template."""

    allowed: bool
    reason: Optional[str] = None
    status_code: int = 200


def _derive_excel_template_title(model: models.Model, method_name: str) -> str:
    """
    Compute a readable fallback title when none is provided.

    Args:
        model: Django model class owning the template.
        method_name: Name of the decorated method.

    Returns:
        Human-readable title.
    """
    base = method_name.replace("_", " ").strip() or "Export"
    base = base[:1].upper() + base[1:]
    verbose_name = getattr(getattr(model, "_meta", None), "verbose_name", None)
    if verbose_name:
        return f"{base} ({verbose_name})"
    return base


def _derive_function_title(func: Callable) -> str:
    """Compute a readable fallback title for function templates."""
    base = getattr(func, "__name__", "").replace("_", " ").strip() or "Excel Export"
    return base[:1].upper() + base[1:]


def _clean_client_value(value: Any) -> str:
    """Normalize client-provided values to bounded strings."""
    try:
        if value is None:
            return ""
        if isinstance(value, (bytes, bytearray)):
            value = value.decode("utf-8", "ignore")
        text = str(value)
    except Exception:
        text = ""
    return text[:1024]


def _extract_client_data(
    request: HttpRequest, template_def: ExcelTemplateDefinition
) -> Dict[str, Any]:
    """
    Extract whitelisted client-provided values from the request (query params only).

    Args:
        request: The HTTP request.
        template_def: The template definition with client_data configuration.

    Returns:
        Dictionary of allowed client data values.
    """
    if not template_def.allow_client_data:
        return {}

    allowed_keys = {str(k) for k in (template_def.client_data_fields or [])}
    if not allowed_keys:
        # If allow_client_data is True but no fields specified, allow all query params
        data: Dict[str, Any] = {}
        for key in request.GET.keys():
            data[key] = _clean_client_value(request.GET.get(key))
        return data

    data = {}
    for key in allowed_keys:
        if key in request.GET:
            data[key] = _clean_client_value(request.GET.get(key))

    return data


class ExcelTemplateRegistry:
    """Keeps track of all registered Excel templates exposed by models."""

    def __init__(self) -> None:
        self._templates: Dict[str, ExcelTemplateDefinition] = {}

    def register(
        self, model: Type[models.Model], method_name: str, meta: ExcelTemplateMeta
    ) -> None:
        """
        Register an Excel template for a model method.

        Args:
            model: Django model class owning the method.
            method_name: Name of the decorated method.
            meta: Raw decorator metadata.
        """
        app_label = model._meta.app_label
        model_name = model._meta.model_name
        url_path = meta.url_path or f"{app_label}/{model_name}/{method_name}"

        merged_config = {**_default_excel_config(), **(meta.config or {})}
        title = meta.title or _derive_excel_template_title(model, method_name)

        definition = ExcelTemplateDefinition(
            model=model,
            method_name=method_name,
            handler=None,
            source="model",
            url_path=url_path,
            config=merged_config,
            roles=tuple(meta.roles or ()),
            permissions=tuple(meta.permissions or ()),
            guard=meta.guard,
            require_authentication=meta.require_authentication,
            title=title,
            allow_client_data=bool(meta.allow_client_data),
            client_data_fields=tuple(meta.client_data_fields or ()),
        )

        self._templates[url_path] = definition
        logger.info(
            "Registered Excel template for %s.%s at /api/%s/%s/<pk>/",
            model.__name__,
            method_name,
            _url_prefix(),
            url_path,
        )

    def register_function(self, func: Callable, meta: ExcelTemplateMeta) -> None:
        """
        Register an Excel template for a standalone function.

        Args:
            func: Callable that returns data for the Excel file.
            meta: Raw decorator metadata.
        """
        module_label = str(getattr(func, "__module__", "")).split(".")[-1] or "excel"
        url_path = meta.url_path or f"{module_label}/{func.__name__}"

        merged_config = {**_default_excel_config(), **(meta.config or {})}
        title = meta.title or _derive_function_title(func)

        definition = ExcelTemplateDefinition(
            model=None,
            method_name=None,
            handler=func,
            source="function",
            url_path=url_path,
            config=merged_config,
            roles=tuple(meta.roles or ()),
            permissions=tuple(meta.permissions or ()),
            guard=meta.guard,
            require_authentication=meta.require_authentication,
            title=title,
            allow_client_data=bool(meta.allow_client_data),
            client_data_fields=tuple(meta.client_data_fields or ()),
        )

        self._templates[url_path] = definition
        logger.info(
            "Registered function Excel template for %s at /api/%s/%s/<pk>/",
            func.__name__,
            _url_prefix(),
            url_path,
        )

    def get(self, url_path: str) -> Optional[ExcelTemplateDefinition]:
        """Retrieve a registered template by its URL path."""
        return self._templates.get(url_path)

    def all(self) -> Dict[str, ExcelTemplateDefinition]:
        """Expose all templates (primarily for introspection and tests)."""
        return dict(self._templates)


excel_template_registry = ExcelTemplateRegistry()


def model_excel_template(
    *,
    url: Optional[str] = None,
    title: Optional[str] = None,
    config: Optional[Dict[str, Any]] = None,
    roles: Optional[Iterable[str]] = None,
    permissions: Optional[Iterable[str]] = None,
    guard: Optional[str] = None,
    require_authentication: bool = True,
    allow_client_data: bool = False,
    client_data_fields: Optional[Iterable[str]] = None,
) -> Callable:
    """
    Decorator to expose a model method as an Excel endpoint.

    The decorated function should return data in one of these formats:
    - Single sheet: list[list[Any]] where first row is headers
    - Multi-sheet: dict[str, list[list[Any]]] where keys are sheet names

    The decorated function can optionally accept a `request` parameter to access
    the HTTP request object, including query parameters and user information.

    Args:
        url: Relative URL (under /api/<prefix>/) for the Excel endpoint. Defaults to
             <app_label>/<model_name>/<function_name>.
        title: Optional human-readable label surfaced to the frontend.
        config: Optional style overrides with keys:
            - header_style: Dict with font, fill, alignment for headers
            - cell_style: Dict for data cells
            - column_widths: Dict mapping column indices to widths, or "auto" for auto-sizing
            - freeze_panes: Boolean to freeze header row
            - number_formats: Dict mapping column indices to Excel number formats
            - sheet_name: Default sheet name (for single-sheet exports)
            - alternating_rows: Dict with enable, even_fill_color, odd_fill_color
            - borders: Dict with enable, color, style
            - auto_filter: Boolean to enable auto-filter on headers
        roles: Optional iterable of RBAC role names required to access the Excel.
        permissions: Optional iterable of Django permission strings required.
        guard: Optional GraphQL guard name (defaults to "retrieve" when omitted).
        require_authentication: Whether authentication is mandatory (default True).
        allow_client_data: When True, query parameters are extracted and available via
            request.rail_excel_client_data. Default False.
        client_data_fields: Optional iterable of allowed query parameter names. If not
            specified and allow_client_data is True, all query params are allowed.

    Returns:
        The original function with attached metadata for automatic registration.

    Example:
        @model_excel_template(
            url="products/export",
            title="Product Export",
            config={
                "header_style": {"bold": True, "fill_color": "4472C4"},
                "freeze_panes": True,
                "column_widths": "auto",
            },
            allow_client_data=True,
            client_data_fields=["category", "status"],
        )
        def export_products(self, request):
            # Access query params
            category = request.GET.get("category")
            # Or via client_data
            client_data = getattr(request, "rail_excel_client_data", {})
            return [
                ["Name", "Price"],
                *[[p.name, p.price] for p in Product.objects.all()]
            ]
    """

    def decorator(func: Callable) -> Callable:
        func._excel_template_meta = ExcelTemplateMeta(
            url_path=url,
            config=config or {},
            roles=tuple(roles or ()),
            permissions=tuple(permissions or ()),
            guard=guard,
            require_authentication=require_authentication,
            title=title,
            allow_client_data=allow_client_data,
            client_data_fields=tuple(client_data_fields or ()),
        )
        return func

    return decorator


def excel_template(
    *,
    url: Optional[str] = None,
    title: Optional[str] = None,
    config: Optional[Dict[str, Any]] = None,
    roles: Optional[Iterable[str]] = None,
    permissions: Optional[Iterable[str]] = None,
    guard: Optional[str] = None,
    require_authentication: bool = True,
    allow_client_data: bool = False,
    client_data_fields: Optional[Iterable[str]] = None,
) -> Callable:
    """
    Decorator to expose a standalone function as an Excel endpoint.

    The decorated function should return data in one of these formats:
    - Single sheet: list[list[Any]] where first row is headers
    - Multi-sheet: dict[str, list[list[Any]]] where keys are sheet names

    The decorated function receives `request` and `pk` parameters.

    Args:
        url: Relative URL (under /api/<prefix>/) for the Excel endpoint. Defaults to
             <module>/<function_name>.
        title: Optional human-readable label surfaced to the frontend.
        config: Optional style overrides (same as model_excel_template).
        roles: Optional iterable of RBAC role names required to access the Excel.
        permissions: Optional iterable of Django permission strings required.
        guard: Optional GraphQL guard name (ignored when no model is associated).
        require_authentication: Whether authentication is mandatory (default True).
        allow_client_data: When True, query parameters are extracted and available via
            request.rail_excel_client_data. Default False.
        client_data_fields: Optional iterable of allowed query parameter names.

    Returns:
        The decorated function.
    """

    def decorator(func: Callable) -> Callable:
        meta = ExcelTemplateMeta(
            url_path=url,
            config=config or {},
            roles=tuple(roles or ()),
            permissions=tuple(permissions or ()),
            guard=guard,
            require_authentication=require_authentication,
            title=title,
            allow_client_data=allow_client_data,
            client_data_fields=tuple(client_data_fields or ()),
        )
        func._excel_template_meta = meta
        excel_template_registry.register_function(func, meta)
        return func

    return decorator


def _register_model_excel_templates(sender: Any, **kwargs: Any) -> None:
    """
    Signal handler to register decorated methods once models are ready.

    Args:
        sender: Model class being prepared.
    """
    if not hasattr(sender, "_meta"):
        return
    if sender._meta.abstract:
        return

    for attr_name, attr in inspect.getmembers(sender, predicate=callable):
        meta: Optional[ExcelTemplateMeta] = getattr(attr, "_excel_template_meta", None)
        if not meta:
            continue

        excel_template_registry.register(sender, attr_name, meta)


class_prepared.connect(
    _register_model_excel_templates, dispatch_uid="excel_template_registration"
)


def _register_existing_models_if_ready() -> None:
    """Register templates for models that were loaded before the module import."""
    try:
        from django.apps import apps

        if not apps.ready:
            return

        for model in apps.get_models():
            _register_model_excel_templates(model)
    except Exception as exc:  # pragma: no cover - defensive during startup
        logger.debug("Skipping eager Excel template registration: %s", exc)


_register_existing_models_if_ready()


def _resolve_request_user(request: HttpRequest):
    """
    Retrieve a user from the request session or Authorization header.

    Args:
        request: The HTTP request.

    Returns:
        The authenticated user or None.
    """
    user = getattr(request, "user", None)
    if user and getattr(user, "is_authenticated", False):
        return user

    if not get_user_from_token:
        return user

    auth_header = request.META.get("HTTP_AUTHORIZATION", "")
    if auth_header and (
        auth_header.startswith("Bearer ") or auth_header.startswith("Token ")
    ):
        parts = auth_header.split(" ", 1)
        if len(parts) == 2:
            token = parts[1].strip()
            if token:
                try:
                    fallback_user = get_user_from_token(token)
                except Exception:  # pragma: no cover - defensive
                    fallback_user = None
                if fallback_user and getattr(fallback_user, "is_authenticated", False):
                    request.user = fallback_user
                    return fallback_user

    return user


def evaluate_excel_template_access(
    template_def: ExcelTemplateDefinition,
    user: Optional[Any],
    *,
    instance: Optional[models.Model] = None,
) -> ExcelTemplateAccessDecision:
    """
    Determine whether a user can access a registered Excel template.

    Args:
        template_def: Template definition entry.
        user: Django user (may be anonymous/None).
        instance: Optional model instance for guard evaluation.

    Returns:
        ExcelTemplateAccessDecision describing the authorization result.
    """
    is_authenticated = bool(user and getattr(user, "is_authenticated", False))

    if template_def.require_authentication and not is_authenticated:
        return ExcelTemplateAccessDecision(
            allowed=False,
            reason="Authentication is required to access this Excel export.",
            status_code=401,
        )

    if not is_authenticated:
        # Anonymous access explicitly allowed; no further checks required.
        return ExcelTemplateAccessDecision(allowed=True)

    if getattr(user, "is_superuser", False):
        return ExcelTemplateAccessDecision(allowed=True)

    required_permissions = tuple(template_def.permissions or ())
    if required_permissions and not any(
        user.has_perm(permission) for permission in required_permissions
    ):
        return ExcelTemplateAccessDecision(
            allowed=False,
            reason="Missing permission to generate this Excel export.",
            status_code=403,
        )

    required_roles = tuple(template_def.roles or ())
    if required_roles:
        if not role_manager:
            logger.warning(
                "Role manager unavailable while enforcing Excel template roles for %s",
                template_def.url_path,
            )
            return ExcelTemplateAccessDecision(
                allowed=False,
                reason="Role control is unavailable.",
                status_code=403,
            )
        try:
            user_roles = set(role_manager.get_user_roles(user))
        except Exception as exc:  # pragma: no cover - defensive logging
            logger.warning("Unable to fetch roles for %s: %s", user, exc)
            user_roles = set()

        if not user_roles.intersection(set(required_roles)):
            return ExcelTemplateAccessDecision(
                allowed=False,
                reason="Required role missing for this Excel export.",
                status_code=403,
            )

    if permission_manager and OperationType and template_def.model:
        try:
            model_label = template_def.model._meta.label_lower
            permission_state = permission_manager.check_operation_permission(
                user, model_label, OperationType.READ
            )
        except Exception as exc:  # pragma: no cover - defensive logging
            logger.warning(
                "Permission manager check failed for %s: %s (denying access)",
                template_def.model.__name__,
                exc,
            )
            return ExcelTemplateAccessDecision(
                allowed=False,
                reason="Permission verification unavailable.",
                status_code=403,
            )
        else:
            if not permission_state.allowed:
                return ExcelTemplateAccessDecision(
                    allowed=False,
                    reason=permission_state.reason or "Access denied for this Excel export.",
                    status_code=403,
                )

    if template_def.model and instance is None:
        return ExcelTemplateAccessDecision(allowed=True)

    guard_name = template_def.guard or ("retrieve" if template_def.model else None)
    if guard_name and template_def.model:
        if not get_model_graphql_meta:
            logger.warning(
                "GraphQL meta unavailable while enforcing Excel template guard '%s' for %s",
                guard_name,
                template_def.url_path,
            )
            return ExcelTemplateAccessDecision(
                allowed=False,
                reason="Access control is unavailable.",
                status_code=403,
            )

        graphql_meta = None
        try:
            graphql_meta = get_model_graphql_meta(template_def.model)
        except Exception as exc:  # pragma: no cover - defensive
            logger.debug(
                "GraphQLMeta unavailable for %s: %s",
                template_def.model.__name__,
                exc,
            )

        if graphql_meta:
            guard_state = None
            try:
                guard_state = graphql_meta.describe_operation_guard(
                    guard_name,
                    user=user,
                    instance=instance,
                )
            except Exception as exc:  # pragma: no cover - defensive
                logger.warning(
                    "Failed to evaluate guard '%s' for %s: %s (denying access)",
                    guard_name,
                    template_def.model.__name__,
                    exc,
                )
                return ExcelTemplateAccessDecision(
                    allowed=False,
                    reason="Operation guard unavailable.",
                    status_code=403,
                )

            if (
                guard_state
                and guard_state.get("guarded")
                and not guard_state.get("allowed", True)
            ):
                return ExcelTemplateAccessDecision(
                    allowed=False,
                    reason=guard_state.get("reason")
                    or "Access denied by operation guard.",
                    status_code=403,
                )

    return ExcelTemplateAccessDecision(allowed=True)


def authorize_excel_template_access(
    request: HttpRequest,
    template_def: ExcelTemplateDefinition,
    instance: Optional[models.Model] = None,
) -> Optional[JsonResponse]:
    """
    Authorize access to an Excel template and return a denial response if not allowed.

    Args:
        request: The HTTP request.
        template_def: Template definition to check access for.
        instance: Optional model instance for guard evaluation.

    Returns:
        JsonResponse with error details if access denied, None if allowed.
    """
    user = _resolve_request_user(request)
    decision = evaluate_excel_template_access(
        template_def,
        user=user,
        instance=instance,
    )
    if decision.allowed:
        return None
    detail = decision.reason or (
        "Authentication is required to access this Excel export."
        if decision.status_code == 401
        else "Access denied for this Excel export."
    )
    return JsonResponse(
        {"error": "Forbidden", "detail": detail}, status=decision.status_code
    )


def _call_model_method(method: Optional[Callable], request: HttpRequest) -> Any:
    """
    Call a model method with optional request parameter.

    Args:
        method: The method to call.
        request: The HTTP request.

    Returns:
        The return value of the method.
    """
    if not callable(method):
        return []
    try:
        signature = inspect.signature(method)
    except (TypeError, ValueError):
        return method()

    params = signature.parameters
    request_param = params.get("request")
    if not request_param:
        for param in params.values():
            if param.annotation is HttpRequest:
                request_param = param
                break
    if request_param:
        if request_param.kind == request_param.KEYWORD_ONLY:
            return method(request=request)
        return method(request)
    has_var_keyword = any(
        param.kind == param.VAR_KEYWORD for param in params.values()
    )
    has_var_positional = any(
        param.kind == param.VAR_POSITIONAL for param in params.values()
    )
    if has_var_keyword:
        return method(request=request)
    if has_var_positional:
        return method(request)
    return method()


def _call_function_handler(
    handler: Callable, request: HttpRequest, pk: Optional[str]
) -> Any:
    """
    Call a standalone function handler with request and pk parameters.

    Args:
        handler: The function to call.
        request: The HTTP request.
        pk: The primary key.

    Returns:
        The return value of the function.
    """
    try:
        signature = inspect.signature(handler)
    except (TypeError, ValueError):
        return handler(request, pk)

    params = signature.parameters
    positional_params = [
        param
        for param in params.values()
        if param.kind in (param.POSITIONAL_ONLY, param.POSITIONAL_OR_KEYWORD)
    ]
    allows_varargs = any(param.kind == param.VAR_POSITIONAL for param in params.values())
    accepts_kwargs = any(param.kind == param.VAR_KEYWORD for param in params.values())
    request_param = params.get("request")
    args: List[Any] = []
    kwargs: Dict[str, Any] = {}

    if request_param:
        if request_param.kind == request_param.KEYWORD_ONLY:
            kwargs["request"] = request
        else:
            args.append(request)
    elif accepts_kwargs:
        kwargs["request"] = request

    if pk is not None:
        if len(args) < len(positional_params) or allows_varargs:
            args.append(pk)
        elif "pk" in params:
            kwargs["pk"] = pk
        elif "id" in params:
            kwargs["id"] = pk
        elif accepts_kwargs:
            kwargs["pk"] = pk

    return handler(*args, **kwargs)


def _get_excel_data(
    request: HttpRequest,
    instance: Optional[models.Model],
    template_def: ExcelTemplateDefinition,
    pk: Optional[str] = None,
) -> ExcelData:
    """
    Get data from the decorated method or function.

    Args:
        request: The HTTP request.
        instance: The model instance (for model methods).
        template_def: The template definition.
        pk: The primary key.

    Returns:
        Data in single-sheet or multi-sheet format.
    """
    if template_def.source == "model":
        method = getattr(instance, template_def.method_name or "", None)
        return _call_model_method(method, request)
    elif template_def.handler:
        return _call_function_handler(template_def.handler, request, pk)
    return []


def _format_cell_value(value: Any, config: Dict[str, Any], col_idx: int) -> Any:
    """
    Format a cell value for Excel.

    Args:
        value: The raw value.
        config: The template configuration.
        col_idx: The column index (0-based).

    Returns:
        The formatted value.
    """
    if value is None:
        return ""

    # Handle datetime/date first
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return value
    if isinstance(value, Decimal):
        return float(value)

    return value


def _get_column_width(value: Any, min_width: int = 8, max_width: int = 50) -> int:
    """
    Calculate column width based on content.

    Args:
        value: The cell value.
        min_width: Minimum column width.
        max_width: Maximum column width.

    Returns:
        The calculated width.
    """
    if value is None:
        return min_width

    str_value = str(value)
    # Account for multi-line content
    if "\n" in str_value:
        lines = str_value.split("\n")
        width = max(len(line) for line in lines)
    else:
        width = len(str_value)

    # Add some padding
    width = width + 2

    return max(min_width, min(width, max_width))


def _apply_header_style(
    cell: Any, header_style: Dict[str, Any], borders: Optional[Dict[str, Any]] = None
) -> None:
    """
    Apply header styling to a cell.

    Args:
        cell: The openpyxl cell.
        header_style: The header style configuration.
        borders: Optional border configuration.
    """
    if not OPENPYXL_AVAILABLE:
        return

    font_kwargs = {
        "bold": header_style.get("bold", True),
        "size": header_style.get("font_size", 11),
    }
    if header_style.get("font_color"):
        font_kwargs["color"] = header_style["font_color"]

    cell.font = Font(**font_kwargs)

    if header_style.get("fill_color"):
        cell.fill = PatternFill(
            start_color=header_style["fill_color"],
            end_color=header_style["fill_color"],
            fill_type="solid",
        )

    alignment = header_style.get("alignment", "center")
    cell.alignment = Alignment(
        horizontal=alignment,
        vertical="center",
        wrap_text=header_style.get("wrap_text", False),
    )

    if borders and borders.get("enable", True):
        border_color = borders.get("color", "D4D4D4")
        border_style = borders.get("style", "thin")
        side = Side(style=border_style, color=border_color)
        cell.border = Border(left=side, right=side, top=side, bottom=side)


def _apply_cell_style(
    cell: Any,
    cell_style: Dict[str, Any],
    row_idx: int,
    alternating: Optional[Dict[str, Any]] = None,
    borders: Optional[Dict[str, Any]] = None,
) -> None:
    """
    Apply cell styling to a data cell.

    Args:
        cell: The openpyxl cell.
        cell_style: The cell style configuration.
        row_idx: The row index (0-based, for alternating colors).
        alternating: Optional alternating row style configuration.
        borders: Optional border configuration.
    """
    if not OPENPYXL_AVAILABLE:
        return

    font_kwargs = {
        "size": cell_style.get("font_size", 11),
    }
    if cell_style.get("bold"):
        font_kwargs["bold"] = True
    if cell_style.get("font_color"):
        font_kwargs["color"] = cell_style["font_color"]

    cell.font = Font(**font_kwargs)

    # Apply alternating row colors
    if alternating and alternating.get("enable", True):
        if row_idx % 2 == 0:
            fill_color = alternating.get("even_fill_color", "F2F2F2")
        else:
            fill_color = alternating.get("odd_fill_color", "FFFFFF")
        if fill_color and fill_color.upper() != "FFFFFF":
            cell.fill = PatternFill(
                start_color=fill_color,
                end_color=fill_color,
                fill_type="solid",
            )

    alignment = cell_style.get("alignment", "left")
    cell.alignment = Alignment(
        horizontal=alignment,
        vertical="center",
        wrap_text=cell_style.get("wrap_text", False),
    )

    if borders and borders.get("enable", True):
        border_color = borders.get("color", "D4D4D4")
        border_style = borders.get("style", "thin")
        side = Side(style=border_style, color=border_color)
        cell.border = Border(left=side, right=side, top=side, bottom=side)


def _apply_number_format(
    cell: Any, col_idx: int, value: Any, config: Dict[str, Any]
) -> None:
    """
    Apply number formatting to a cell.

    Args:
        cell: The openpyxl cell.
        col_idx: The column index (0-based).
        value: The cell value.
        config: The template configuration.
    """
    number_formats = config.get("number_formats", {})

    # Check for explicit format for this column
    if col_idx in number_formats:
        cell.number_format = number_formats[col_idx]
    elif str(col_idx) in number_formats:
        cell.number_format = number_formats[str(col_idx)]
    # Apply default formats based on type
    elif isinstance(value, datetime):
        cell.number_format = config.get("datetime_format", "YYYY-MM-DD HH:MM:SS")
    elif isinstance(value, date):
        cell.number_format = config.get("date_format", "YYYY-MM-DD")
    elif isinstance(value, (Decimal, float)) and not isinstance(value, bool):
        cell.number_format = config.get("decimal_format", "#,##0.00")


def _calculate_column_widths(
    sheet_data: ExcelSheetData, explicit_widths: Union[str, Dict[int, int], None]
) -> Dict[int, int]:
    """
    Calculate column widths for a sheet.

    Args:
        sheet_data: The sheet data.
        explicit_widths: Explicit widths or "auto" for auto-sizing.

    Returns:
        Dictionary mapping column index to width.
    """
    if isinstance(explicit_widths, dict):
        return explicit_widths

    # Auto-calculate widths
    widths: Dict[int, int] = {}
    for row in sheet_data:
        for col_idx, value in enumerate(row):
            current_width = widths.get(col_idx, 8)
            calculated_width = _get_column_width(value)
            widths[col_idx] = max(current_width, calculated_width)

    return widths


def render_excel_sheet(
    worksheet: Any,
    sheet_data: ExcelSheetData,
    config: Dict[str, Any],
) -> None:
    """
    Render data to an Excel worksheet with styling.

    Args:
        worksheet: The openpyxl worksheet.
        sheet_data: The data to render (first row is headers).
        config: The template configuration.
    """
    if not OPENPYXL_AVAILABLE or not sheet_data:
        return

    header_style = config.get("header_style", DEFAULT_HEADER_STYLE)
    cell_style = config.get("cell_style", DEFAULT_CELL_STYLE)
    alternating = config.get("alternating_rows", DEFAULT_ALTERNATING_ROW_STYLE)
    borders = config.get("borders", DEFAULT_BORDER_STYLE)
    column_widths = config.get("column_widths", "auto")

    # Calculate column widths
    calculated_widths = _calculate_column_widths(sheet_data, column_widths)

    # Write data
    for row_idx, row_data in enumerate(sheet_data):
        for col_idx, value in enumerate(row_data):
            cell = worksheet.cell(row=row_idx + 1, column=col_idx + 1)
            formatted_value = _format_cell_value(value, config, col_idx)
            cell.value = formatted_value

            if row_idx == 0:
                # Header row
                _apply_header_style(cell, header_style, borders)
            else:
                # Data rows
                _apply_cell_style(cell, cell_style, row_idx - 1, alternating, borders)

            # Apply number format
            _apply_number_format(cell, col_idx, value, config)

    # Apply column widths
    for col_idx, width in calculated_widths.items():
        col_letter = get_column_letter(col_idx + 1)
        worksheet.column_dimensions[col_letter].width = width

    # Freeze panes
    if config.get("freeze_panes", True) and len(sheet_data) > 0:
        worksheet.freeze_panes = "A2"

    # Auto-filter
    if config.get("auto_filter", True) and len(sheet_data) > 0:
        num_cols = max(len(row) for row in sheet_data) if sheet_data else 0
        if num_cols > 0:
            last_col_letter = get_column_letter(num_cols)
            worksheet.auto_filter.ref = f"A1:{last_col_letter}1"


def render_excel(
    data: ExcelData,
    config: Optional[Dict[str, Any]] = None,
) -> bytes:
    """
    Render data to an Excel file.

    Args:
        data: Single sheet data (list[list]) or multi-sheet data (dict[str, list[list]]).
        config: Optional style configuration.

    Returns:
        Excel file as bytes.

    Raises:
        RuntimeError: If openpyxl is not installed.
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is required for Excel export")

    config = {**_default_excel_config(), **(config or {})}

    workbook = Workbook()

    # Determine if single-sheet or multi-sheet
    if isinstance(data, dict):
        # Multi-sheet format
        # Remove the default sheet
        workbook.remove(workbook.active)

        for sheet_name, sheet_data in data.items():
            worksheet = workbook.create_sheet(title=str(sheet_name)[:31])  # Excel limit
            render_excel_sheet(worksheet, sheet_data, config)
    else:
        # Single-sheet format
        worksheet = workbook.active
        worksheet.title = str(config.get("sheet_name", "Sheet1"))[:31]
        render_excel_sheet(worksheet, data, config)

    # Save to bytes
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def _sanitize_filename(filename: str) -> str:
    """
    Sanitize a filename for safe use in Content-Disposition header.

    Args:
        filename: The raw filename.

    Returns:
        The sanitized filename.
    """
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", filename).strip("._")
    return cleaned or "export"


def _hash_payload(payload: Dict[str, Any]) -> str:
    """
    Create a hash of a payload for cache key generation.

    Args:
        payload: The payload to hash.

    Returns:
        The hash string.
    """
    serialized = json.dumps(payload, sort_keys=True, default=str).encode("utf-8")
    return hashlib.sha256(serialized).hexdigest()


def _cache_settings_for_template(template_def: ExcelTemplateDefinition) -> Dict[str, Any]:
    """
    Get cache settings for a template, merging global and template-specific config.

    Args:
        template_def: The template definition.

    Returns:
        Merged cache settings.
    """
    overrides = {}
    if isinstance(template_def.config, dict):
        overrides = template_def.config.get("cache") or {}
    return _merge_dict(_excel_cache(), overrides)


def _build_excel_cache_key(
    template_def: ExcelTemplateDefinition,
    *,
    pk: Optional[str],
    user: Optional[Any],
    cache_settings: Dict[str, Any],
) -> Optional[str]:
    """
    Build a cache key for an Excel export.

    Args:
        template_def: The template definition.
        pk: The primary key.
        user: The user.
        cache_settings: Cache settings.

    Returns:
        The cache key or None if caching is disabled.
    """
    if not cache_settings.get("enable", False):
        return None

    payload: Dict[str, Any] = {"template": template_def.url_path, "pk": pk}
    if cache_settings.get("vary_on_user", True):
        payload["user"] = getattr(user, "id", None) or "anon"

    key_prefix = cache_settings.get("key_prefix", "rail:excel_cache")
    return f"{key_prefix}:{_hash_payload(payload)}"


def _excel_job_cache_key(job_id: str) -> str:
    """Get cache key for an async Excel job."""
    return f"rail:excel_job:{job_id}"


def _excel_job_payload_key(job_id: str) -> str:
    """Get cache key for an async Excel job payload."""
    return f"rail:excel_job_payload:{job_id}"


def _get_excel_storage_dir(async_settings: Dict[str, Any]) -> Path:
    """
    Get the storage directory for async Excel files.

    Args:
        async_settings: Async settings.

    Returns:
        The storage directory path.
    """
    storage_dir = async_settings.get("storage_dir")
    if storage_dir:
        path = Path(str(storage_dir))
    elif getattr(settings, "MEDIA_ROOT", None):
        path = Path(settings.MEDIA_ROOT) / "rail_excel"
    else:
        path = Path(tempfile.gettempdir()) / "rail_excel"
    path.mkdir(parents=True, exist_ok=True)
    return path


def _get_excel_job(job_id: str) -> Optional[Dict[str, Any]]:
    """Get an async Excel job from cache."""
    return cache.get(_excel_job_cache_key(job_id))


def _set_excel_job(job_id: str, job: Dict[str, Any], *, timeout: int) -> None:
    """Set an async Excel job in cache."""
    cache.set(_excel_job_cache_key(job_id), job, timeout=timeout)


def _update_excel_job(
    job_id: str, updates: Dict[str, Any], *, timeout: int
) -> Optional[Dict[str, Any]]:
    """Update an async Excel job in cache."""
    job = _get_excel_job(job_id)
    if not job:
        return None
    job.update(updates)
    _set_excel_job(job_id, job, timeout=timeout)
    return job


def _delete_excel_job(job_id: str) -> None:
    """Delete an async Excel job from cache."""
    cache.delete(_excel_job_cache_key(job_id))
    cache.delete(_excel_job_payload_key(job_id))


def _parse_iso_datetime(value: Optional[str]) -> Optional[datetime]:
    """Parse an ISO datetime string."""
    if not value:
        return None
    try:
        return datetime.fromisoformat(value)
    except ValueError:
        return None


def _job_access_allowed(request: Any, job: Dict[str, Any]) -> bool:
    """Check if the requesting user can access a job."""
    user = _resolve_request_user(request)
    if not user or not getattr(user, "is_authenticated", False):
        return False
    if getattr(user, "is_superuser", False):
        return True
    owner_id = job.get("owner_id")
    return bool(owner_id and str(owner_id) == str(getattr(user, "id", "")))


def _notify_excel_job_webhook(
    job: Dict[str, Any], async_settings: Dict[str, Any]
) -> None:
    """Send a webhook notification for a completed Excel job."""
    webhook_url = async_settings.get("webhook_url")
    if not webhook_url:
        return
    try:
        import requests
    except Exception:
        logger.warning("requests is unavailable; cannot post Excel webhook")
        return
    headers = async_settings.get("webhook_headers") or {}
    timeout = int(async_settings.get("webhook_timeout_seconds", 10))
    try:
        requests.post(webhook_url, json=job, headers=headers, timeout=timeout)
    except Exception as exc:  # pragma: no cover - network dependent
        logger.warning("Failed to notify Excel webhook: %s", exc)


def _build_job_request(owner_id: Optional[Any]) -> HttpRequest:
    """Build a mock request for async job processing."""
    request = HttpRequest()
    user = None
    if owner_id:
        try:
            user_model = get_user_model()
            user = user_model.objects.filter(pk=owner_id).first()
        except Exception:
            user = None
    request.user = user or AnonymousUser()
    return request


def _run_excel_job(job_id: str) -> None:
    """Run an async Excel generation job."""
    job = _get_excel_job(job_id)
    if not job:
        return

    async_settings = _excel_async()
    timeout = int(async_settings.get("expires_seconds", 3600))
    payload = cache.get(_excel_job_payload_key(job_id))
    if not payload:
        _update_excel_job(
            job_id, {"status": "failed", "error": "Missing job payload"}, timeout=timeout
        )
        return

    _update_excel_job(
        job_id,
        {"status": "running", "started_at": timezone.now().isoformat()},
        timeout=timeout,
    )

    template_path = payload.get("template_path")
    template_def = excel_template_registry.get(str(template_path))
    if not template_def:
        _update_excel_job(
            job_id, {"status": "failed", "error": "Template not found"}, timeout=timeout
        )
        return

    pk = payload.get("pk")
    instance: Optional[models.Model] = None
    if template_def.model:
        try:
            instance = template_def.model.objects.get(pk=pk)
        except (template_def.model.DoesNotExist, ValidationError, ValueError, TypeError):
            _update_excel_job(
                job_id,
                {"status": "failed", "error": "Instance not found"},
                timeout=timeout,
            )
            return

    request = _build_job_request(job.get("owner_id"))

    try:
        data = _get_excel_data(
            request, instance, template_def, pk=str(pk) if pk else None
        )
        excel_bytes = render_excel(data, config=template_def.config)
    except Exception as exc:  # pragma: no cover - defensive
        logger.exception("Async Excel job failed: %s", exc)
        _update_excel_job(
            job_id,
            {"status": "failed", "error": str(exc) if _excel_expose_errors() else "Excel render failed"},
            timeout=timeout,
        )
        _notify_excel_job_webhook(_get_excel_job(job_id) or job, async_settings)
        return

    storage_dir = _get_excel_storage_dir(async_settings)
    filename = payload.get("filename") or template_def.url_path.replace("/", "-")
    filename = _sanitize_filename(filename)
    file_path = storage_dir / f"{job_id}.xlsx"
    try:
        with open(file_path, "wb") as handle:
            handle.write(excel_bytes)
    except OSError as exc:
        _update_excel_job(
            job_id,
            {"status": "failed", "error": f"Failed to persist Excel: {exc}"},
            timeout=timeout,
        )
        _notify_excel_job_webhook(_get_excel_job(job_id) or job, async_settings)
        return

    _update_excel_job(
        job_id,
        {
            "status": "completed",
            "completed_at": timezone.now().isoformat(),
            "file_path": str(file_path),
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "filename": f"{filename}.xlsx",
        },
        timeout=timeout,
    )
    _notify_excel_job_webhook(_get_excel_job(job_id) or job, async_settings)


try:
    from celery import shared_task
except Exception:  # pragma: no cover - optional dependency
    shared_task = None

if shared_task:
    @shared_task(name="rail_django.excel_job")
    def excel_job_task(job_id: str) -> None:
        _run_excel_job(job_id)
else:
    excel_job_task = None


def generate_excel_async(
    *,
    request: HttpRequest,
    template_def: ExcelTemplateDefinition,
    pk: Optional[str],
) -> Dict[str, Any]:
    """
    Generate an Excel file asynchronously.

    Args:
        request: The HTTP request.
        template_def: The template definition.
        pk: The primary key.

    Returns:
        Job information including job_id and status URLs.
    """
    async_settings = _excel_async()
    backend = str(async_settings.get("backend", "thread")).lower()
    expires_seconds = int(async_settings.get("expires_seconds", 3600))

    job_id = str(uuid.uuid4())
    now = timezone.now()
    owner = _resolve_request_user(request)
    job = {
        "id": job_id,
        "status": "pending",
        "created_at": now.isoformat(),
        "expires_at": (now + timedelta(seconds=expires_seconds)).isoformat(),
        "owner_id": getattr(owner, "id", None),
    }

    payload = {
        "template_path": template_def.url_path,
        "pk": pk,
        "filename": template_def.title,
    }

    _set_excel_job(job_id, job, timeout=expires_seconds)
    cache.set(_excel_job_payload_key(job_id), payload, timeout=expires_seconds)

    if backend == "thread":
        thread = threading.Thread(target=_run_excel_job, args=(job_id,), daemon=True)
        thread.start()
    elif backend == "celery":
        if not excel_job_task:
            _update_excel_job(
                job_id,
                {"status": "failed", "error": "Celery is not available"},
                timeout=expires_seconds,
            )
            raise RuntimeError("Celery backend not available")
        excel_job_task.delay(job_id)
    elif backend == "rq":
        try:
            import django_rq
        except Exception as exc:
            _update_excel_job(
                job_id,
                {"status": "failed", "error": "RQ is not available"},
                timeout=expires_seconds,
            )
            raise RuntimeError("RQ backend not available") from exc
        queue_name = async_settings.get("queue", "default")
        queue = django_rq.get_queue(queue_name)
        queue.enqueue(_run_excel_job, job_id)
    else:
        _update_excel_job(
            job_id,
            {"status": "failed", "error": "Unknown async backend"},
            timeout=expires_seconds,
        )
        raise RuntimeError("Unknown async backend")

    status_path = f"/api/{_url_prefix().rstrip('/')}/jobs/{job_id}/"
    download_path = f"/api/{_url_prefix().rstrip('/')}/jobs/{job_id}/download/"
    return {
        "job_id": job_id,
        "status": "pending",
        "status_url": request.build_absolute_uri(status_path),
        "download_url": request.build_absolute_uri(download_path),
        "expires_in": expires_seconds,
    }


def _cleanup_excel_job_files(job: Dict[str, Any]) -> None:
    """Clean up files from an async Excel job."""
    file_path = job.get("file_path")
    if not file_path:
        return
    try:
        Path(str(file_path)).unlink(missing_ok=True)
    except Exception:
        return


@method_decorator(csrf_exempt, name="dispatch")
@method_decorator(
    jwt_required if jwt_required else (lambda view: view), name="dispatch"
)
class ExcelTemplateView(View):
    """Serve model Excel files rendered with openpyxl."""

    http_method_names = ["get"]

    def get(
        self,
        request: HttpRequest,
        template_path: str,
        *args: Any,
        **kwargs: Any,
    ) -> HttpResponse:
        """
        Generate an Excel file for a given model instance.

        Args:
            request: Incoming Django request.
            template_path: Relative template path registered for the model.

        Query Parameters:
            pk: Primary key of the model instance (optional for function templates).

        Returns:
            Excel response or JSON error when unavailable.
        """
        # Get pk from query parameters
        pk = request.GET.get("pk")

        template_def = excel_template_registry.get(template_path)
        if not template_def:
            self._log_template_event(
                request,
                success=False,
                error_message="Template not found",
                template_path=template_path,
                pk=pk,
            )
            return JsonResponse(
                {"error": "Template not found", "template": template_path}, status=404
            )

        if not OPENPYXL_AVAILABLE:
            self._log_template_event(
                request,
                success=False,
                error_message="openpyxl not available",
                template_def=template_def,
                template_path=template_path,
                pk=pk,
            )
            return JsonResponse(
                {
                    "error": "Excel export unavailable",
                    "detail": "openpyxl is not installed" if _excel_expose_errors() else None,
                },
                status=500,
            )

        rate_limit_response = self._check_rate_limit(request, template_def)
        if rate_limit_response:
            return rate_limit_response

        instance: Optional[models.Model] = None
        if template_def.model:
            if not pk:
                return JsonResponse(
                    {"error": "Missing required parameter 'pk'"}, status=400
                )
            try:
                instance = template_def.model.objects.get(pk=pk)
            except template_def.model.DoesNotExist:
                self._log_template_event(
                    request,
                    success=False,
                    error_message="Instance not found",
                    template_def=template_def,
                    template_path=template_path,
                    pk=pk,
                )
                return JsonResponse(
                    {
                        "error": "Instance not found",
                        "model": template_def.model._meta.label,
                        "pk": pk,
                    },
                    status=404,
                )
            except (ValidationError, ValueError, TypeError):
                self._log_template_event(
                    request,
                    success=False,
                    error_message="Invalid primary key",
                    template_def=template_def,
                    template_path=template_path,
                    pk=pk,
                )
                return JsonResponse(
                    {"error": "Invalid primary key", "pk": pk}, status=400
                )

        denial = self._authorize_template_access(request, template_def, instance)
        if denial:
            self._log_template_event(
                request,
                success=False,
                error_message="Forbidden",
                template_def=template_def,
                template_path=template_path,
                pk=pk,
            )
            return denial

        # Extract client data and attach to request
        client_data = _extract_client_data(request, template_def)
        setattr(request, "rail_excel_client_data", client_data)

        if self._parse_async_request(request):
            async_settings = _excel_async()
            if not async_settings.get("enable", False):
                return JsonResponse(
                    {"error": "Async Excel jobs are disabled"}, status=400
                )
            try:
                job_payload = generate_excel_async(
                    request=request,
                    template_def=template_def,
                    pk=pk,
                )
            except Exception as exc:
                return JsonResponse(
                    {
                        "error": "Failed to enqueue Excel job",
                        "detail": str(exc) if _excel_expose_errors() else None,
                    },
                    status=500,
                )
            self._log_template_event(
                request,
                success=True,
                template_def=template_def,
                template_path=template_path,
                pk=pk,
            )
            return JsonResponse(job_payload, status=202)

        cache_settings = _cache_settings_for_template(template_def)
        cache_key = _build_excel_cache_key(
            template_def,
            pk=pk,
            user=self._resolve_request_user(request),
            cache_settings=cache_settings,
        )
        if cache_key:
            cached_excel = cache.get(cache_key)
            if cached_excel:
                response = HttpResponse(
                    cached_excel,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                filename = self._resolve_filename(template_def, pk)
                response["Content-Disposition"] = f'attachment; filename="{filename}"'
                self._log_template_event(
                    request,
                    success=True,
                    template_def=template_def,
                    template_path=template_path,
                    pk=pk,
                )
                return response

        try:
            data = _get_excel_data(request, instance, template_def, pk)
            excel_bytes = render_excel(data, config=template_def.config)
        except Exception as exc:  # pragma: no cover - defensive logging branch
            model_name = (
                template_def.model.__name__
                if template_def.model
                else template_def.url_path
            )
            logger.exception(
                "Failed to render Excel for %s pk=%s: %s",
                model_name,
                pk,
                exc,
            )
            self._log_template_event(
                request,
                success=False,
                error_message=str(exc),
                template_def=template_def,
                template_path=template_path,
                pk=pk,
            )
            detail = str(exc) if _excel_expose_errors() else "Failed to render Excel"
            return JsonResponse(
                {"error": "Failed to render Excel", "detail": detail}, status=500
            )

        if cache_key:
            cache_timeout = int(cache_settings.get("timeout_seconds", 300))
            cache.set(cache_key, excel_bytes, timeout=cache_timeout)

        filename = self._resolve_filename(template_def, pk)
        response = HttpResponse(
            excel_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        self._log_template_event(
            request,
            success=True,
            template_def=template_def,
            template_path=template_path,
            pk=pk,
        )
        return response

    def _authorize_template_access(
        self,
        request: HttpRequest,
        template_def: ExcelTemplateDefinition,
        instance: Optional[models.Model],
    ) -> Optional[JsonResponse]:
        """Apply RBAC and permission requirements before rendering the Excel."""
        return authorize_excel_template_access(request, template_def, instance)

    def _parse_async_request(self, request: HttpRequest) -> bool:
        """Check if the request wants async processing."""
        value = request.GET.get("async")
        if value is None:
            return False
        return str(value).strip().lower() in {"1", "true", "yes", "on"}

    def _resolve_filename(self, template_def: ExcelTemplateDefinition, pk: Optional[str]) -> str:
        """Generate a filename for the Excel download."""
        if template_def.model:
            base_name = f"{template_def.model._meta.model_name}-{pk}"
        else:
            base_name = f"{template_def.url_path.replace('/', '-')}-{pk}"
        return f"{_sanitize_filename(base_name)}.xlsx"

    def _get_rate_limit_identifier(
        self, request: HttpRequest, rate_limit: Dict[str, Any]
    ) -> str:
        """Get the rate limit identifier for a request."""
        user = _resolve_request_user(request)
        if user and getattr(user, "is_authenticated", False):
            return f"user:{user.id}"

        trusted_proxies = rate_limit.get("trusted_proxies") or []
        remote_addr = request.META.get("REMOTE_ADDR", "")
        ip_address = remote_addr or "unknown"

        if self._is_trusted_proxy(remote_addr, trusted_proxies):
            forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR", "")
            if forwarded_for:
                ip_address = forwarded_for.split(",")[0].strip()

        return f"ip:{ip_address}"

    def _is_trusted_proxy(self, remote_addr: str, trusted_proxies: Iterable[str]) -> bool:
        """Check if a remote address is a trusted proxy."""
        if not remote_addr:
            return False
        for proxy in trusted_proxies:
            proxy = str(proxy).strip()
            if not proxy:
                continue
            if "/" in proxy:
                try:
                    if ipaddress.ip_address(remote_addr) in ipaddress.ip_network(
                        proxy, strict=False
                    ):
                        return True
                except ValueError:
                    continue
            if remote_addr == proxy:
                return True
        return False

    def _check_rate_limit(
        self, request: HttpRequest, template_def: ExcelTemplateDefinition
    ) -> Optional[JsonResponse]:
        """Check rate limits for the request."""
        config = _excel_rate_limit()
        overrides = template_def.config.get("rate_limit") or {}
        config = _merge_dict(config, overrides)
        if not config.get("enable", True):
            return None

        window_seconds = int(config.get("window_seconds", 60))
        max_requests = int(config.get("max_requests", 30))
        identifier = self._get_rate_limit_identifier(request, config)
        cache_key = f"rail:excel_rl:{identifier}:{template_def.url_path}"

        # Atomic increment with race-condition handling
        try:
            current_count = cache.incr(cache_key)
        except ValueError:
            # Key doesn't exist, try to add it atomically
            if cache.add(cache_key, 1, timeout=window_seconds):
                return None
            # Another request beat us to it, try increment again
            try:
                current_count = cache.incr(cache_key)
            except ValueError:
                # Key expired between add and incr, start fresh
                cache.set(cache_key, 1, timeout=window_seconds)
                return None

        if current_count > max_requests:
            return JsonResponse(
                {"error": "Rate limit exceeded", "retry_after": window_seconds},
                status=429,
            )

        return None

    def _resolve_request_user(self, request: HttpRequest):
        """Retrieve a user from the request session or Authorization header."""
        return _resolve_request_user(request)

    def _log_template_event(
        self,
        request: HttpRequest,
        *,
        success: bool,
        error_message: Optional[str] = None,
        template_def: Optional[ExcelTemplateDefinition] = None,
        template_path: Optional[str] = None,
        pk: Optional[str] = None,
    ) -> None:
        """Log an audit event for template rendering."""
        if not log_audit_event or not AuditEventType:
            return

        details = {
            "action": "excel_template_render",
            "template_path": template_path,
            "pk": pk,
        }
        if template_def:
            if template_def.model:
                details["model"] = template_def.model._meta.label
            details["title"] = template_def.title
            details["source"] = template_def.source

        log_audit_event(
            request,
            AuditEventType.DATA_ACCESS,
            success=success,
            error_message=error_message,
            additional_data=details,
        )


@method_decorator(csrf_exempt, name="dispatch")
@method_decorator(
    jwt_required if jwt_required else (lambda view: view), name="dispatch"
)
class ExcelTemplateCatalogView(View):
    """Expose Excel template catalog metadata for UI-driven workflows."""

    http_method_names = ["get"]

    def get(self, request: HttpRequest) -> JsonResponse:
        """
        Return the catalog of available Excel templates.

        Args:
            request: The HTTP request.

        Returns:
            JSON response with template catalog.
        """
        catalog_settings = _excel_catalog()
        if not catalog_settings.get("enable", True):
            raise Http404("Catalog disabled")

        user = _resolve_request_user(request)
        if catalog_settings.get("require_authentication", True) and not (
            user and getattr(user, "is_authenticated", False)
        ):
            return JsonResponse(
                {"error": "Authentication required"}, status=401
            )

        include_config = bool(catalog_settings.get("include_config", False))
        include_permissions = bool(catalog_settings.get("include_permissions", True))
        filter_by_access = bool(catalog_settings.get("filter_by_access", True))

        templates = []
        for url_path, template_def in sorted(excel_template_registry.all().items()):
            access = evaluate_excel_template_access(template_def, user=user, instance=None)
            if filter_by_access and not access.allowed:
                continue

            entry = {
                "url_path": url_path,
                "title": template_def.title,
                "source": template_def.source,
                "model": template_def.model._meta.label if template_def.model else None,
                "require_authentication": template_def.require_authentication,
            }
            if include_permissions:
                entry.update(
                    {
                        "roles": template_def.roles,
                        "permissions": template_def.permissions,
                        "guard": template_def.guard,
                    }
                )
            if include_config:
                entry["config"] = template_def.config
            if not filter_by_access:
                entry["access"] = {
                    "allowed": access.allowed,
                    "reason": access.reason,
                    "status_code": access.status_code,
                }
            templates.append(entry)

        return JsonResponse({"templates": templates})


@method_decorator(csrf_exempt, name="dispatch")
@method_decorator(
    jwt_required if jwt_required else (lambda view: view), name="dispatch"
)
class ExcelTemplateJobStatusView(View):
    """Return status for async Excel jobs."""

    http_method_names = ["get"]

    def get(self, request: HttpRequest, job_id: str) -> JsonResponse:
        """
        Get the status of an async Excel job.

        Args:
            request: The HTTP request.
            job_id: The job ID.

        Returns:
            JSON response with job status.
        """
        job = _get_excel_job(str(job_id))
        if not job:
            raise Http404("Excel job not found")

        # Check expiration before access check
        expires_at = _parse_iso_datetime(job.get("expires_at"))
        if expires_at and expires_at <= timezone.now():
            _cleanup_excel_job_files(job)
            _delete_excel_job(str(job_id))
            raise Http404("Excel job not found")

        if not _job_access_allowed(request, job):
            return JsonResponse({"error": "Excel job not permitted"}, status=403)

        payload = {
            "job_id": job.get("id"),
            "status": job.get("status"),
            "error": job.get("error"),
            "created_at": job.get("created_at"),
            "completed_at": job.get("completed_at"),
            "expires_at": job.get("expires_at"),
        }
        if job.get("status") == "completed":
            download_path = f"/api/{_url_prefix().rstrip('/')}/jobs/{job_id}/download/"
            payload["download_url"] = request.build_absolute_uri(download_path)

        return JsonResponse(payload)


@method_decorator(csrf_exempt, name="dispatch")
@method_decorator(
    jwt_required if jwt_required else (lambda view: view), name="dispatch"
)
class ExcelTemplateJobDownloadView(View):
    """Download completed Excel job files."""

    http_method_names = ["get"]

    def get(self, request: HttpRequest, job_id: str) -> HttpResponse:
        """
        Download a completed async Excel file.

        Args:
            request: The HTTP request.
            job_id: The job ID.

        Returns:
            File response with the Excel file.
        """
        job = _get_excel_job(str(job_id))
        if not job:
            raise Http404("Excel job not found")

        # Check expiration before access check
        expires_at = _parse_iso_datetime(job.get("expires_at"))
        if expires_at and expires_at <= timezone.now():
            _cleanup_excel_job_files(job)
            _delete_excel_job(str(job_id))
            raise Http404("Excel job not found")

        if not _job_access_allowed(request, job):
            return JsonResponse({"error": "Excel job not permitted"}, status=403)

        if job.get("status") != "completed":
            return JsonResponse({"error": "Excel job not completed"}, status=409)

        file_path = job.get("file_path")
        if not file_path or not Path(str(file_path)).exists():
            return JsonResponse({"error": "Excel job file missing"}, status=410)

        filename = _sanitize_filename(str(job.get("filename") or "export"))
        response = FileResponse(
            open(file_path, "rb"),
            content_type=job.get("content_type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


def excel_urlpatterns():
    """
    Return URL patterns to expose Excel template endpoints under the configured prefix.

    The final URL shape is:
        /api/<prefix>/<template_path>/
        /api/<prefix>/<template_path>/?pk=<pk>

    Where <template_path> defaults to <app_label>/<model_name>/<function_name>.
    The pk parameter is passed as a query parameter.

    Returns:
        List of URL patterns for Excel endpoints.
    """
    prefix = _url_prefix().rstrip("/")
    return [
        path(
            f"{prefix}/catalog/",
            ExcelTemplateCatalogView.as_view(),
            name="excel_template_catalog",
        ),
        path(
            f"{prefix}/jobs/<uuid:job_id>/",
            ExcelTemplateJobStatusView.as_view(),
            name="excel_template_job_status",
        ),
        path(
            f"{prefix}/jobs/<uuid:job_id>/download/",
            ExcelTemplateJobDownloadView.as_view(),
            name="excel_template_job_download",
        ),
        path(
            f"{prefix}/<path:template_path>/",
            ExcelTemplateView.as_view(),
            name="excel_template",
        ),
    ]


__all__ = [
    # Views
    "ExcelTemplateView",
    "ExcelTemplateCatalogView",
    "ExcelTemplateJobStatusView",
    "ExcelTemplateJobDownloadView",
    # Decorators
    "model_excel_template",
    "excel_template",
    # Registry
    "excel_template_registry",
    "ExcelTemplateRegistry",
    "ExcelTemplateDefinition",
    "ExcelTemplateMeta",
    # Access control
    "evaluate_excel_template_access",
    "authorize_excel_template_access",
    "ExcelTemplateAccessDecision",
    # Rendering
    "render_excel",
    "render_excel_sheet",
    # URL patterns
    "excel_urlpatterns",
    # Async
    "generate_excel_async",
    # Constants
    "OPENPYXL_AVAILABLE",
]
