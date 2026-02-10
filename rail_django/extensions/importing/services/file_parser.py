"""CSV/XLSX parser with import limits and header validation."""

from __future__ import annotations

import csv
import io
from collections.abc import Mapping
from typing import Any

from ..constants import ImportIssueCode
from ..types import ParsedImportFile
from .errors import ImportServiceError

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


def _read_uploaded_file(uploaded_file: Any) -> bytes:
    if uploaded_file is None:
        raise ImportServiceError(
            ImportIssueCode.INVALID_FILE_FORMAT,
            "Missing uploaded file.",
        )
    if isinstance(uploaded_file, str):
        return uploaded_file.encode("utf-8")
    if isinstance(uploaded_file, (bytes, bytearray)):
        return bytes(uploaded_file)
    if isinstance(uploaded_file, Mapping):
        for key in ("file", "originFileObj", "raw", "value"):
            nested = uploaded_file.get(key)
            if hasattr(nested, "read"):
                uploaded_file = nested
                break
        else:
            raise ImportServiceError(
                ImportIssueCode.INVALID_FILE_FORMAT,
                "Uploaded file payload is invalid. Expected a binary uploaded file.",
            )
    if hasattr(uploaded_file, "seek"):
        uploaded_file.seek(0)
    if not hasattr(uploaded_file, "read"):
        raise ImportServiceError(
            ImportIssueCode.INVALID_FILE_FORMAT,
            "Uploaded file payload is invalid. Expected a readable file object.",
        )
    content = uploaded_file.read()
    if isinstance(content, str):
        return content.encode("utf-8")
    return bytes(content or b"")


def _validate_size(content: bytes, max_file_size_bytes: int) -> None:
    if len(content) > max_file_size_bytes:
        raise ImportServiceError(
            ImportIssueCode.FILE_TOO_LARGE,
            f"Uploaded file exceeds {max_file_size_bytes} bytes.",
        )


def _parse_csv(content: bytes, *, max_rows: int) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    headers = [str(header).strip() for header in (reader.fieldnames or []) if header]
    if not headers:
        raise ImportServiceError(
            ImportIssueCode.MISSING_REQUIRED_COLUMN,
            "CSV header row is missing.",
        )

    rows: list[dict[str, Any]] = []
    for row in reader:
        if len(rows) >= max_rows:
            raise ImportServiceError(
                ImportIssueCode.ROW_LIMIT_EXCEEDED,
                f"File exceeds row limit of {max_rows}.",
            )
        normalized: dict[str, Any] = {}
        for key in headers:
            normalized[key] = row.get(key)
        if any(value not in (None, "") for value in normalized.values()):
            rows.append(normalized)
    return headers, rows


def _parse_xlsx(content: bytes, *, max_rows: int) -> tuple[list[str], list[dict[str, Any]]]:
    if load_workbook is None:
        raise ImportServiceError(
            ImportIssueCode.INVALID_FILE_FORMAT,
            "XLSX parsing is unavailable because openpyxl is not installed.",
        )

    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    indexed_headers: list[tuple[int, str]] = []
    for column_index, header_value in enumerate(header_row or []):
        if header_value is None:
            continue
        header_name = str(header_value).strip()
        if not header_name:
            continue
        indexed_headers.append((column_index, header_name))
    if not indexed_headers:
        raise ImportServiceError(
            ImportIssueCode.MISSING_REQUIRED_COLUMN,
            "XLSX header row is missing.",
        )
    headers = [header_name for _column_index, header_name in indexed_headers]

    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if len(rows) >= max_rows:
            raise ImportServiceError(
                ImportIssueCode.ROW_LIMIT_EXCEEDED,
                f"File exceeds row limit of {max_rows}.",
            )
        row = {
            header_name: values[column_index] if column_index < len(values) else None
            for column_index, header_name in indexed_headers
        }
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    return headers, rows


def parse_uploaded_file(
    uploaded_file: Any,
    *,
    file_format: str,
    max_rows: int,
    max_file_size_bytes: int,
) -> ParsedImportFile:
    content = _read_uploaded_file(uploaded_file)
    _validate_size(content, max_file_size_bytes=max_file_size_bytes)

    normalized_format = str(file_format).upper()
    if normalized_format == "CSV":
        headers, rows = _parse_csv(content, max_rows=max_rows)
    elif normalized_format == "XLSX":
        headers, rows = _parse_xlsx(content, max_rows=max_rows)
    else:
        raise ImportServiceError(
            ImportIssueCode.INVALID_FILE_FORMAT,
            f"Unsupported file format '{file_format}'.",
        )

    return ParsedImportFile(
        headers=headers,
        rows=rows,
        file_format=normalized_format,
        file_name=getattr(uploaded_file, "name", "upload"),
        file_size_bytes=len(content),
    )
