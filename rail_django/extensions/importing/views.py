"""HTTP views for import template downloads."""

from __future__ import annotations

import csv
import io
from typing import Any

from django.apps import apps
from django.db import models
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.csrf import csrf_exempt

from ..excel.builder import OPENPYXL_AVAILABLE, render_excel

try:
    from ..auth.decorators import jwt_required
except ImportError:  # pragma: no cover
    jwt_required = None

from .services import require_import_access, resolve_template_descriptor

try:  # pragma: no cover - optional dependency
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except Exception:  # pragma: no cover
    load_workbook = None
    get_column_letter = None
    DataValidation = None


TEMPLATE_PREFILLED_ROWS = 5
CSV_PREFILLED_ROWS = 1
FK_CHOICE_SEPARATOR = " | "
FK_LABEL_ATTRS = ("desc", "description", "label", "name", "title")


def _parse_requested_fields(request: HttpRequest) -> list[str]:
    selected: list[str] = []
    seen: set[str] = set()
    for raw_value in request.GET.getlist("fields"):
        for token in str(raw_value).split(","):
            candidate = token.strip()
            if not candidate or candidate in seen:
                continue
            seen.add(candidate)
            selected.append(candidate)
    return selected


def _ordered_columns(descriptor) -> list[dict[str, Any]]:
    ordered: list[dict[str, Any]] = []
    seen: set[str] = set()
    for column in descriptor["required_columns"] + descriptor["optional_columns"]:
        name = str(column["name"]).strip()
        if not name or name in seen:
            continue
        seen.add(name)
        ordered.append(column)
    return ordered


def _select_columns(descriptor, requested_fields: list[str]) -> list[dict[str, Any]]:
    ordered = _ordered_columns(descriptor)
    if not requested_fields:
        return ordered

    requested = set(requested_fields)
    required = {
        str(column["name"]).strip()
        for column in descriptor["required_columns"]
        if str(column["name"]).strip()
    }
    selected = [
        column
        for column in ordered
        if str(column["name"]).strip() in requested
        or str(column["name"]).strip() in required
    ]
    if selected:
        return selected
    return [column for column in ordered if str(column["name"]).strip() in required]


def _template_cell_value(column: dict[str, Any]) -> Any:
    value = column.get("default_value")
    if value is None:
        return ""
    if isinstance(value, (bool, int, float, str)):
        return value
    return str(value)


def _build_template_rows(
    columns: list[dict[str, Any]], row_count: int
) -> list[list[Any]]:
    headers = [str(column["name"]).strip() for column in columns]
    default_row = [_template_cell_value(column) for column in columns]
    rows = [headers]
    for _ in range(max(0, row_count)):
        rows.append(list(default_row))
    return rows


def _describe_related_instance(instance: Any) -> str:
    for attr in FK_LABEL_ATTRS:
        if not hasattr(instance, attr):
            continue
        value = getattr(instance, attr)
        if value not in (None, ""):
            return str(value)
    text = str(instance)
    return text if text not in (None, "") else ""


def _is_fk_field(field: Any) -> bool:
    return bool(
        getattr(field, "is_relation", False)
        and getattr(field, "many_to_one", False)
        and not getattr(field, "many_to_many", False)
    )


def _foreign_key_choices(field: models.Field) -> list[str]:
    if not _is_fk_field(field):
        return []

    remote_field = getattr(field, "remote_field", None)
    related_model = getattr(remote_field, "model", None)
    if related_model is None:
        return []

    pk_name = related_model._meta.pk.name
    queryset = related_model._default_manager.all().order_by(pk_name)
    choices: list[str] = []
    for instance in queryset[:1000]:
        raw_id = getattr(instance, pk_name, None)
        if raw_id in (None, ""):
            continue
        identifier = str(raw_id)
        label = _describe_related_instance(instance)
        if label and label != identifier:
            choices.append(f"{identifier}{FK_CHOICE_SEPARATOR}{label}")
        else:
            choices.append(identifier)
    return choices


def _apply_fk_validations(
    *,
    workbook,
    worksheet,
    model,
    columns: list[dict[str, Any]],
    start_row: int,
    end_row: int,
) -> None:
    if DataValidation is None or get_column_letter is None:
        return
    if end_row < start_row:
        return

    model_fields = {field.name: field for field in model._meta.fields}
    fk_columns: list[tuple[int, models.Field, list[str]]] = []
    for column_index, column in enumerate(columns, start=1):
        field_name = str(column["name"]).strip()
        field = model_fields.get(field_name)
        if field is None or not _is_fk_field(field):
            continue
        choices = _foreign_key_choices(field)
        if choices:
            fk_columns.append((column_index, field, choices))

    if not fk_columns:
        return

    choices_sheet_name = "_choices"
    if choices_sheet_name in workbook.sheetnames:
        workbook.remove(workbook[choices_sheet_name])
    choices_sheet = workbook.create_sheet(title=choices_sheet_name)
    choices_sheet.sheet_state = "hidden"

    for choice_col_index, (column_index, field, choices) in enumerate(
        fk_columns, start=1
    ):
        choice_letter = get_column_letter(choice_col_index)
        choices_sheet.cell(row=1, column=choice_col_index, value=str(field.name))
        for row_index, choice in enumerate(choices, start=2):
            choices_sheet.cell(row=row_index, column=choice_col_index, value=choice)

        last_choice_row = len(choices) + 1
        formula = f"'{choices_sheet_name}'!${choice_letter}$2:${choice_letter}${last_choice_row}"
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.showInputMessage = True
        validation.showErrorMessage = True
        field_label = str(getattr(field, "verbose_name", field.name) or field.name)
        validation.promptTitle = f"Select {field_label}"
        preview = ", ".join(choices[:8])
        if len(choices) > 8:
            preview = f"{preview}, ..."
        validation.prompt = (
            f"Choose an existing value ({FK_CHOICE_SEPARATOR.join(['id', 'label'])}): {preview}"
        )[:255]
        worksheet.add_data_validation(validation)

        target_letter = get_column_letter(column_index)
        validation.add(f"${target_letter}${start_row}:${target_letter}${end_row}")


def _render_xlsx_template(
    *,
    app_label: str,
    model_name: str,
    columns: list[dict[str, Any]],
) -> bytes:
    rows = _build_template_rows(columns, row_count=TEMPLATE_PREFILLED_ROWS)
    content = render_excel(rows)
    if load_workbook is None:
        return content

    workbook = load_workbook(io.BytesIO(content))
    worksheet = workbook.active
    model = apps.get_model(app_label, model_name)
    _apply_fk_validations(
        workbook=workbook,
        worksheet=worksheet,
        model=model,
        columns=columns,
        start_row=2,
        end_row=TEMPLATE_PREFILLED_ROWS + 1,
    )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


@method_decorator(csrf_exempt, name="dispatch")
@method_decorator(
    jwt_required if jwt_required else (lambda view: view), name="dispatch"
)
class ModelImportTemplateDownloadView(View):
    """Download a template file for model imports."""

    http_method_names = ["get"]

    def get(
        self, request: HttpRequest, app_label: str, model_name: str, *args, **kwargs
    ) -> HttpResponse:
        user = getattr(request, "user", None)
        require_import_access(user, app_label=app_label, model_name=model_name)

        try:
            descriptor = resolve_template_descriptor(
                app_label=app_label, model_name=model_name
            )
        except LookupError:
            return JsonResponse(
                {"error": f"Model '{app_label}.{model_name}' not found."},
                status=404,
            )

        requested_fields = _parse_requested_fields(request)
        selected_columns = _select_columns(descriptor, requested_fields)

        requested_format = str(request.GET.get("format", "csv")).strip().lower()
        if requested_format not in {"csv", "xlsx"}:
            return JsonResponse(
                {"error": "Unsupported format. Use 'csv' or 'xlsx'."},
                status=400,
            )

        file_base = f"{app_label}-{model_name.lower()}-import-template"
        if requested_format == "xlsx":
            if not OPENPYXL_AVAILABLE:
                return JsonResponse(
                    {
                        "error": "XLSX export unavailable because openpyxl is not installed."
                    },
                    status=500,
                )
            content = _render_xlsx_template(
                app_label=app_label,
                model_name=model_name,
                columns=selected_columns,
            )
            response = HttpResponse(
                content,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{file_base}.xlsx"'
            return response

        rows = _build_template_rows(selected_columns, row_count=CSV_PREFILLED_ROWS)
        buffer = io.StringIO(newline="")
        writer = csv.writer(buffer)
        writer.writerows(rows)
        response = HttpResponse(
            buffer.getvalue().encode("utf-8-sig"),
            content_type="text/csv; charset=utf-8",
        )
        response["Content-Disposition"] = f'attachment; filename="{file_base}.csv"'
        return response
