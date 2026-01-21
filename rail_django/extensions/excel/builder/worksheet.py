"""
Excel worksheet rendering utilities.

This module provides functions for rendering data to Excel worksheets,
including column width calculation and sheet-level formatting.
"""

from typing import Any, Dict, List, Union

# Optional openpyxl support
try:
    from openpyxl.utils import get_column_letter

    OPENPYXL_UTILS_AVAILABLE = True
except ImportError:  # pragma: no cover - optional dependency
    get_column_letter = None  # type: ignore
    OPENPYXL_UTILS_AVAILABLE = False

from ..config import (
    DEFAULT_ALTERNATING_ROW_STYLE,
    DEFAULT_BORDER_STYLE,
    DEFAULT_CELL_STYLE,
    DEFAULT_HEADER_STYLE,
    ExcelSheetData,
)
from .styles import (
    _apply_cell_style,
    _apply_header_style,
    _apply_number_format,
    _format_cell_value,
    _get_column_width,
)


def _calculate_column_widths(
    sheet_data: ExcelSheetData, explicit_widths: Union[str, Dict[int, int], None]
) -> Dict[int, int]:
    """
    Calculate column widths for a sheet.

    Args:
        sheet_data: The sheet data.
        explicit_widths: Explicit widths or "auto" for auto-sizing.

    Returns:
        Dictionary mapping column index to width.
    """
    if isinstance(explicit_widths, dict):
        return explicit_widths

    # Auto-calculate widths
    widths: Dict[int, int] = {}
    for row in sheet_data:
        for col_idx, value in enumerate(row):
            current_width = widths.get(col_idx, 8)
            calculated_width = _get_column_width(value)
            widths[col_idx] = max(current_width, calculated_width)

    return widths


def render_excel_sheet(
    worksheet: Any,
    sheet_data: ExcelSheetData,
    config: Dict[str, Any],
) -> None:
    """
    Render data to an Excel worksheet with styling.

    Args:
        worksheet: The openpyxl worksheet.
        sheet_data: The data to render (first row is headers).
        config: The template configuration.
    """
    if not OPENPYXL_UTILS_AVAILABLE or not sheet_data:
        return

    header_style = config.get("header_style", DEFAULT_HEADER_STYLE)
    cell_style = config.get("cell_style", DEFAULT_CELL_STYLE)
    alternating = config.get("alternating_rows", DEFAULT_ALTERNATING_ROW_STYLE)
    borders = config.get("borders", DEFAULT_BORDER_STYLE)
    column_widths = config.get("column_widths", "auto")

    # Calculate column widths
    calculated_widths = _calculate_column_widths(sheet_data, column_widths)

    # Write data
    for row_idx, row_data in enumerate(sheet_data):
        for col_idx, value in enumerate(row_data):
            cell = worksheet.cell(row=row_idx + 1, column=col_idx + 1)
            formatted_value = _format_cell_value(value, config, col_idx)
            cell.value = formatted_value

            if row_idx == 0:
                # Header row
                _apply_header_style(cell, header_style, borders)
            else:
                # Data rows
                _apply_cell_style(cell, cell_style, row_idx - 1, alternating, borders)

            # Apply number format
            _apply_number_format(cell, col_idx, value, config)

    # Apply column widths
    for col_idx, width in calculated_widths.items():
        col_letter = get_column_letter(col_idx + 1)
        worksheet.column_dimensions[col_letter].width = width

    # Freeze panes
    if config.get("freeze_panes", True) and len(sheet_data) > 0:
        worksheet.freeze_panes = "A2"

    # Auto-filter
    if config.get("auto_filter", True) and len(sheet_data) > 0:
        num_cols = max(len(row) for row in sheet_data) if sheet_data else 0
        if num_cols > 0:
            last_col_letter = get_column_letter(num_cols)
            worksheet.auto_filter.ref = f"A1:{last_col_letter}1"


__all__ = [
    "OPENPYXL_UTILS_AVAILABLE",
    "_calculate_column_widths",
    "render_excel_sheet",
]
