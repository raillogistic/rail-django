"""
Excel cell styling and formatting utilities.

This module provides functions for applying styles to Excel cells,
including headers, data cells, borders, and number formats.
"""

from datetime import date, datetime
from decimal import Decimal
from typing import Any, Dict, Optional

# Optional openpyxl support
try:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    OPENPYXL_STYLES_AVAILABLE = True
except ImportError:  # pragma: no cover - optional dependency
    Alignment = None  # type: ignore
    Border = None  # type: ignore
    Font = None  # type: ignore
    PatternFill = None  # type: ignore
    Side = None  # type: ignore
    OPENPYXL_STYLES_AVAILABLE = False

from ..config import (
    DEFAULT_ALTERNATING_ROW_STYLE,
    DEFAULT_BORDER_STYLE,
    DEFAULT_CELL_STYLE,
    DEFAULT_HEADER_STYLE,
)


def _format_cell_value(value: Any, config: Dict[str, Any], col_idx: int) -> Any:
    """
    Format a cell value for Excel.

    Args:
        value: The raw value.
        config: The template configuration.
        col_idx: The column index (0-based).

    Returns:
        The formatted value.
    """
    if value is None:
        return ""

    # Handle datetime/date first
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return value
    if isinstance(value, Decimal):
        return float(value)

    return value


def _get_column_width(value: Any, min_width: int = 8, max_width: int = 50) -> int:
    """
    Calculate column width based on content.

    Args:
        value: The cell value.
        min_width: Minimum column width.
        max_width: Maximum column width.

    Returns:
        The calculated width.
    """
    if value is None:
        return min_width

    str_value = str(value)
    # Account for multi-line content
    if "\n" in str_value:
        lines = str_value.split("\n")
        width = max(len(line) for line in lines)
    else:
        width = len(str_value)

    # Add some padding
    width = width + 2

    return max(min_width, min(width, max_width))


def _apply_header_style(
    cell: Any, header_style: Dict[str, Any], borders: Optional[Dict[str, Any]] = None
) -> None:
    """
    Apply header styling to a cell.

    Args:
        cell: The openpyxl cell.
        header_style: The header style configuration.
        borders: Optional border configuration.
    """
    if not OPENPYXL_STYLES_AVAILABLE:
        return

    font_kwargs: Dict[str, Any] = {
        "bold": header_style.get("bold", True),
        "size": header_style.get("font_size", 11),
    }
    if header_style.get("font_color"):
        font_kwargs["color"] = header_style["font_color"]

    cell.font = Font(**font_kwargs)

    if header_style.get("fill_color"):
        cell.fill = PatternFill(
            start_color=header_style["fill_color"],
            end_color=header_style["fill_color"],
            fill_type="solid",
        )

    alignment = header_style.get("alignment", "center")
    cell.alignment = Alignment(
        horizontal=alignment,
        vertical="center",
        wrap_text=header_style.get("wrap_text", False),
    )

    if borders and borders.get("enable", True):
        border_color = borders.get("color", "D4D4D4")
        border_style = borders.get("style", "thin")
        side = Side(style=border_style, color=border_color)
        cell.border = Border(left=side, right=side, top=side, bottom=side)


def _apply_cell_style(
    cell: Any,
    cell_style: Dict[str, Any],
    row_idx: int,
    alternating: Optional[Dict[str, Any]] = None,
    borders: Optional[Dict[str, Any]] = None,
) -> None:
    """
    Apply cell styling to a data cell.

    Args:
        cell: The openpyxl cell.
        cell_style: The cell style configuration.
        row_idx: The row index (0-based, for alternating colors).
        alternating: Optional alternating row style configuration.
        borders: Optional border configuration.
    """
    if not OPENPYXL_STYLES_AVAILABLE:
        return

    font_kwargs: Dict[str, Any] = {
        "size": cell_style.get("font_size", 11),
    }
    if cell_style.get("bold"):
        font_kwargs["bold"] = True
    if cell_style.get("font_color"):
        font_kwargs["color"] = cell_style["font_color"]

    cell.font = Font(**font_kwargs)

    # Apply alternating row colors
    if alternating and alternating.get("enable", True):
        if row_idx % 2 == 0:
            fill_color = alternating.get("even_fill_color", "F2F2F2")
        else:
            fill_color = alternating.get("odd_fill_color", "FFFFFF")
        if fill_color and fill_color.upper() != "FFFFFF":
            cell.fill = PatternFill(
                start_color=fill_color,
                end_color=fill_color,
                fill_type="solid",
            )

    alignment = cell_style.get("alignment", "left")
    cell.alignment = Alignment(
        horizontal=alignment,
        vertical="center",
        wrap_text=cell_style.get("wrap_text", False),
    )

    if borders and borders.get("enable", True):
        border_color = borders.get("color", "D4D4D4")
        border_style = borders.get("style", "thin")
        side = Side(style=border_style, color=border_color)
        cell.border = Border(left=side, right=side, top=side, bottom=side)


def _apply_number_format(
    cell: Any, col_idx: int, value: Any, config: Dict[str, Any]
) -> None:
    """
    Apply number formatting to a cell.

    Args:
        cell: The openpyxl cell.
        col_idx: The column index (0-based).
        value: The cell value.
        config: The template configuration.
    """
    number_formats = config.get("number_formats", {})

    # Check for explicit format for this column
    if col_idx in number_formats:
        cell.number_format = number_formats[col_idx]
    elif str(col_idx) in number_formats:
        cell.number_format = number_formats[str(col_idx)]
    # Apply default formats based on type
    elif isinstance(value, datetime):
        cell.number_format = config.get("datetime_format", "YYYY-MM-DD HH:MM:SS")
    elif isinstance(value, date):
        cell.number_format = config.get("date_format", "YYYY-MM-DD")
    elif isinstance(value, (Decimal, float)) and not isinstance(value, bool):
        cell.number_format = config.get("decimal_format", "#,##0.00")


__all__ = [
    "OPENPYXL_STYLES_AVAILABLE",
    "_format_cell_value",
    "_get_column_width",
    "_apply_header_style",
    "_apply_cell_style",
    "_apply_number_format",
]
