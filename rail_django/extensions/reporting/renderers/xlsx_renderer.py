"""
Excel (XLSX) export renderer.

Requires ``openpyxl`` as an optional dependency. This renderer is only
registered if openpyxl is installed (``pip install rail-django[xlsx]``).

Attributes:
    XlsxRenderer: Renderer producing XLSX output.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any, Optional

from .base import ExportRenderer

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class XlsxRenderer(ExportRenderer):
    """
    Renderer Excel (XLSX) pour les exports de données reporting.

    Produit un classeur Excel formaté avec en-têtes stylisés, largeur
    automatique des colonnes, et support des types natifs (nombres, dates).

    Nécessite ``openpyxl``. Non disponible si le package n'est pas installé.

    Options disponibles:
        - ``sheet_name`` (str): Nom de la feuille (défaut: ``Données``).
        - ``freeze_header`` (bool): Figer la ligne d'en-tête (défaut: ``True``).
        - ``auto_width`` (bool): Ajuster la largeur des colonnes (défaut: ``True``).
        - ``columns`` (list[str]): Colonnes à inclure et leur ordre.

    Raises:
        ImportError: Si ``openpyxl`` n'est pas installé.
    """

    format_name = "xlsx"
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    file_extension = "xlsx"

    def __init__(self) -> None:
        if not HAS_OPENPYXL:
            raise ImportError(
                "Le renderer XLSX necessite 'openpyxl'. "
                "Installez-le avec: pip install openpyxl"
            )

    def render(
        self,
        payload: dict[str, Any],
        *,
        options: Optional[dict[str, Any]] = None,
    ) -> bytes:
        """
        Transforme un payload reporting en fichier Excel.

        Args:
            payload: Dictionnaire contenant ``rows`` et ``columns``.
            options: Options XLSX (sheet_name, freeze_header, auto_width, columns).

        Returns:
            Contenu XLSX en bytes.
        """
        options = options or {}
        sheet_name = str(options.get("sheet_name", "Données"))
        freeze_header = bool(options.get("freeze_header", True))
        auto_width = bool(options.get("auto_width", True))

        rows = payload.get("rows") or []
        if not rows:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            output = BytesIO()
            wb.save(output)
            return output.getvalue()

        # Determine column order
        column_names = options.get("columns")
        if not column_names:
            columns_meta = payload.get("columns") or []
            if columns_meta:
                column_names = [col.get("name", "") for col in columns_meta if col.get("name")]
            else:
                column_names = list(rows[0].keys()) if rows else []

        # Column labels
        columns_meta = payload.get("columns") or []
        label_map = {
            col.get("name", ""): col.get("label", col.get("name", ""))
            for col in columns_meta
        }

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        # Write header
        for col_idx, name in enumerate(column_names, 1):
            cell = ws.cell(row=1, column=col_idx, value=label_map.get(name, name))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Write data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, name in enumerate(column_names, 1):
                value = row.get(name)
                if isinstance(value, (list, dict)):
                    import json
                    value = json.dumps(value, ensure_ascii=False, default=str)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-width columns
        if auto_width:
            for col_idx, name in enumerate(column_names, 1):
                header_len = len(label_map.get(name, name))
                max_len = header_len
                for row in rows[:100]:  # Sample first 100 rows for perf
                    val = row.get(name)
                    if val is not None:
                        max_len = max(max_len, len(str(val)))
                adjusted_width = min(max_len + 4, 60)
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Freeze header row
        if freeze_header:
            ws.freeze_panes = "A2"

        output = BytesIO()
        wb.save(output)
        return output.getvalue()


__all__ = ["XlsxRenderer"]
