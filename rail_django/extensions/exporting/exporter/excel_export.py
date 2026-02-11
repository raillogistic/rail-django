"""Excel Export Functionality

This module provides Excel export functionality as a mixin class.
"""

import io
from typing import Any, Callable, List, Optional, Union

from ..exceptions import ExportError

# Optional Excel support
try:
    import openpyxl
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ExcelExportMixin:
    """Mixin providing Excel export functionality."""

    def export_to_excel(
        self,
        fields: list[Union[str, dict[str, str]]],
        variables: Optional[dict[str, Any]] = None,
        ordering: Optional[Union[str, list[str]]] = None,
        max_rows: Optional[int] = None,
        parsed_fields: Optional[list[dict[str, str]]] = None,
        output: Optional[io.BytesIO] = None,
        progress_callback: Optional[Callable[[int], None]] = None,
        *,
        presets: Optional[List[str]] = None,
        distinct_on: Optional[List[str]] = None,
        group_by: Optional[str] = None,
    ) -> bytes:
        """Export model data to Excel format with professional styling.

        Args:
            fields: List of field definitions (string or dict format).
            variables: Filter variables.
            ordering: Ordering expression(s).
            max_rows: Optional max rows cap.
            parsed_fields: Pre-validated field configurations.
            output: Optional output BytesIO.
            progress_callback: Callback for progress updates.
            presets: Optional list of preset names.
            distinct_on: Optional list of field names for DISTINCT ON.
            group_by: Optional field accessor used to group rows (xlsx only).

        Returns:
            Excel file content as bytes.

        Raises:
            ExportError: If openpyxl is not available.
        """
        if not EXCEL_AVAILABLE:
            raise ExportError(
                "Excel export requires openpyxl package. "
                "Install with: pip install openpyxl"
            )

        group_by_accessor = self.validate_group_by(
            group_by, export_settings=self.export_settings
        )

        # Use non-write-only mode for better styling support
        write_only = bool(self.export_settings.get("excel_write_only", False))
        if group_by_accessor:
            # Grouped export relies on richer styling support.
            write_only = False
        workbook = openpyxl.Workbook(write_only=write_only)
        worksheet = workbook.active if not write_only else workbook.create_sheet()
        worksheet.title = f"{self.model_name} Export"

        # Hide gridlines (only works in non-write-only mode)
        if not write_only:
            worksheet.sheet_view.showGridLines = False

        # Professional style definitions
        header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        header_fill = PatternFill(
            start_color="2F5496", end_color="2F5496", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        # Data row styles
        data_font = Font(size=10, name="Calibri")
        data_alignment = Alignment(vertical="center", wrap_text=False)

        # Row number column style
        row_num_font = Font(size=10, name="Calibri", color="666666")
        row_num_alignment = Alignment(horizontal="center", vertical="center")

        # Group header styles
        group_font = Font(size=10, name="Calibri", bold=True, color="1F3763")
        group_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        group_fill_odd = PatternFill(
            start_color="DCE6F7", end_color="DCE6F7", fill_type="solid"
        )
        group_fill_even = PatternFill(
            start_color="EAF0FB", end_color="EAF0FB", fill_type="solid"
        )

        # Alternating row colors
        even_row_fill = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        )
        odd_row_fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )
        row_num_fill = PatternFill(
            start_color="F8F9FC", end_color="F8F9FC", fill_type="solid"
        )
        thin_side = Side(style="thin", color="D9D9D9")
        medium_side = Side(style="medium", color="9DB4D8")
        cell_border = Border(
            left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
        )
        group_border = Border(
            left=medium_side, right=medium_side, top=medium_side, bottom=medium_side
        )

        # Parse field configurations
        if parsed_fields is None:
            parsed_fields = self.validate_fields(
                fields, export_settings=self.export_settings
            )

        # Write headers - first column is "#" for row numbers
        headers = ["#"] + [parsed_field["title"] for parsed_field in parsed_fields]
        if write_only:
            header_row = []
            for header in headers:
                cell = WriteOnlyCell(worksheet, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = cell_border
                header_row.append(cell)
            worksheet.append(header_row)
        else:
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = cell_border

        # Write data rows
        query_fields = [field["accessor"] for field in parsed_fields]
        if group_by_accessor and group_by_accessor not in query_fields:
            query_fields.append(group_by_accessor)
        queryset = self.get_queryset(
            variables,
            ordering,
            fields=query_fields,
            max_rows=max_rows,
            presets=presets,
            skip_validation=True,  # Already validated at view level
            distinct_on=distinct_on,
        )

        processed = 0
        progress_every = int(
            (self.export_settings.get("async_jobs") or {}).get(
                "progress_update_rows", 500
            )
        )
        if progress_every <= 0:
            progress_every = 500

        # Track max width per column for auto-sizing (including # column)
        column_widths = [3] + [len(str(h)) for h in headers[1:]]

        row_counter = 0  # Sequential row number for # column
        sheet_row = 2

        def write_group_header(title: str, fill: PatternFill) -> None:
            nonlocal sheet_row
            if write_only:
                row = []
                for col_index in range(len(headers)):
                    value = title if col_index == 0 else ""
                    cell = WriteOnlyCell(worksheet, value=value)
                    cell.font = group_font
                    cell.fill = fill
                    cell.alignment = group_alignment
                    cell.border = group_border
                    row.append(cell)
                worksheet.append(row)
            else:
                end_col = len(headers)
                for col_num in range(1, end_col + 1):
                    value = title if col_num == 1 else ""
                    cell = worksheet.cell(row=sheet_row, column=col_num, value=value)
                    cell.font = group_font
                    cell.fill = fill
                    cell.alignment = group_alignment
                    cell.border = group_border
                merge_range = f"A{sheet_row}:{get_column_letter(end_col)}{sheet_row}"
                worksheet.merge_cells(merge_range)
                worksheet.row_dimensions[sheet_row].height = 22
            sheet_row += 1

        def write_data_row(instance: Any, row_fill: PatternFill) -> None:
            nonlocal row_counter, sheet_row, processed
            row_counter += 1
            if write_only:
                row = []
                num_cell = WriteOnlyCell(worksheet, value=row_counter)
                num_cell.font = row_num_font
                num_cell.fill = row_num_fill
                num_cell.alignment = row_num_alignment
                num_cell.border = cell_border
                row.append(num_cell)
                for col_idx, parsed_field in enumerate(parsed_fields):
                    accessor = parsed_field["accessor"]
                    value = self.get_field_value(instance, accessor)
                    cell = WriteOnlyCell(worksheet, value=value)
                    cell.font = data_font
                    cell.fill = row_fill
                    cell.alignment = data_alignment
                    cell.border = cell_border
                    row.append(cell)
                    val_len = len(str(value)) if value else 0
                    if val_len > column_widths[col_idx + 1]:
                        column_widths[col_idx + 1] = min(val_len, 50)
                worksheet.append(row)
            else:
                num_cell = worksheet.cell(row=sheet_row, column=1, value=row_counter)
                num_cell.font = row_num_font
                num_cell.fill = row_num_fill
                num_cell.alignment = row_num_alignment
                num_cell.border = cell_border
                for col_num, parsed_field in enumerate(parsed_fields, 2):
                    accessor = parsed_field["accessor"]
                    value = self.get_field_value(instance, accessor)
                    cell = worksheet.cell(row=sheet_row, column=col_num, value=value)
                    cell.font = data_font
                    cell.fill = row_fill
                    cell.alignment = data_alignment
                    cell.border = cell_border
                    val_len = len(str(value)) if value else 0
                    if val_len > column_widths[col_num - 1]:
                        column_widths[col_num - 1] = min(val_len, 50)
            sheet_row += 1
            processed += 1
            if progress_callback and processed % progress_every == 0:
                progress_callback(processed)

        if group_by_accessor:
            groups: list[dict[str, Any]] = []
            group_index: dict[str, int] = {}
            for instance in queryset.iterator():
                group_value = self.get_field_value(instance, group_by_accessor)
                group_label = (
                    str(group_value)
                    if group_value not in (None, "")
                    else "Non renseigne"
                )
                group_key = group_label
                index = group_index.get(group_key)
                if index is None:
                    group_index[group_key] = len(groups)
                    groups.append({"label": group_label, "rows": [instance]})
                else:
                    groups[index]["rows"].append(instance)

            group_title = self.get_field_verbose_name(group_by_accessor)
            for idx, group in enumerate(groups):
                header_fill = group_fill_even if idx % 2 == 0 else group_fill_odd
                write_group_header(
                    f"{group_title}: {group['label']} ({len(group['rows'])} lignes)",
                    header_fill,
                )
                for instance in group["rows"]:
                    row_fill = even_row_fill if (row_counter + 1) % 2 == 0 else odd_row_fill
                    write_data_row(instance, row_fill)
        else:
            for instance in queryset.iterator():
                row_fill = even_row_fill if (row_counter + 1) % 2 == 0 else odd_row_fill
                write_data_row(instance, row_fill)

        # Apply column widths and additional formatting (non write-only only)
        if not write_only:
            # Set column widths with padding
            for col_idx, width in enumerate(column_widths, 1):
                column_letter = get_column_letter(col_idx)
                if col_idx == 1:  # # column - fixed narrow width
                    worksheet.column_dimensions[column_letter].width = 6
                else:
                    worksheet.column_dimensions[column_letter].width = min(
                        width + 3, 50
                    )

            # Set row height for header
            worksheet.row_dimensions[1].height = 25

        # Save to bytes
        output = output or io.BytesIO()
        workbook.save(output)
        return output.getvalue() if isinstance(output, io.BytesIO) else b""
