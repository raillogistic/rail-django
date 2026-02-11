"""
Unit tests for export helpers and ModelExporter behavior.
"""

import csv
import io
from decimal import Decimal

import pytest
from django.test import TestCase

from rail_django.extensions.exporting import (
    EXCEL_AVAILABLE,
    ExportError,
    ModelExporter,
    _sanitize_filename,
)
from tests.models import TestCompany, TestCustomer

pytestmark = pytest.mark.unit


class TestExporting(TestCase):
    def setUp(self):
        self.customer_safe = TestCustomer.objects.create(
            nom_client="Alpha",
            prenom_client="Alice",
            email_client="alpha@example.com",
            telephone_client="",
            adresse_client="",
            ville_client="",
            code_postal="",
            pays_client="",
            solde_compte=Decimal("10.50"),
            est_actif=True,
        )
        self.customer_formula = TestCustomer.objects.create(
            nom_client="=SUM(1,1)",
            prenom_client="Bob",
            email_client="test@example.com",
            telephone_client="",
            adresse_client="",
            ville_client="",
            code_postal="",
            pays_client="",
            solde_compte=Decimal("20.00"),
            est_actif=False,
        )

    def _settings(self, fields, model_label="tests.testcustomer", **overrides):
        settings = {
            "export_fields": {model_label: list(fields)},
            "require_export_fields": True,
            "require_field_permissions": False,
        }
        settings.update(overrides)
        return settings

    def test_sanitize_filename_falls_back_when_empty(self):
        self.assertEqual(_sanitize_filename("////"), "export")

    def test_validate_fields_requires_allowlist(self):
        exporter = ModelExporter(
            "tests",
            "TestCustomer",
            export_settings={"require_export_fields": True},
        )
        with self.assertRaises(ExportError):
            exporter.validate_fields(["nom_client"])

    def test_validate_fields_blocks_sensitive_field(self):
        export_settings = self._settings(
            ["email_client"], sensitive_fields=["email_client"]
        )
        exporter = ModelExporter(
            "tests", "TestCustomer", export_settings=export_settings
        )
        with self.assertRaises(ExportError) as context:
            exporter.validate_fields(["email_client"])
        self.assertIn("email_client", str(context.exception))

    def test_validate_fields_accepts_allowlisted(self):
        export_settings = self._settings(["nom_client", "email_client"])
        exporter = ModelExporter(
            "tests", "TestCustomer", export_settings=export_settings
        )
        parsed = exporter.validate_fields(["nom_client"])
        self.assertEqual(parsed[0]["accessor"], "nom_client")
        self.assertTrue(parsed[0]["title"])

    def test_export_to_csv_sanitizes_formula_and_formats_bool(self):
        export_settings = self._settings(["nom_client", "est_actif"])
        exporter = ModelExporter(
            "tests", "TestCustomer", export_settings=export_settings
        )
        csv_data = exporter.export_to_csv(
            ["nom_client", "est_actif"], ordering=["nom_client"]
        )

        rows = list(csv.reader(io.StringIO(csv_data)))
        data_rows = rows[1:]
        names = {row[0] for row in data_rows}
        statuses = {row[1] for row in data_rows}

        self.assertIn("'=SUM(1,1)", names)
        self.assertEqual(statuses, {"✅", "❌"})

    def test_field_formatter_mask(self):
        export_settings = self._settings(
            ["email_client"],
            field_formatters={
                "tests.testcustomer": {
                    "email_client": {"type": "mask", "show_last": 4}
                }
            },
        )
        exporter = ModelExporter(
            "tests", "TestCustomer", export_settings=export_settings
        )
        masked = exporter.get_field_value(self.customer_safe, "email_client")
        raw = self.customer_safe.email_client
        expected = "*" * (len(raw) - 4) + raw[-4:]
        self.assertEqual(masked, expected)

    def test_validate_filters_rejects_disallowed_keys(self):
        export_settings = self._settings(
            ["nom_client"],
            filterable_fields={"tests.testcustomer": ["nom_client"]},
            allowed_filter_lookups=["icontains"],
            allowed_filter_transforms=[],
        )
        exporter = ModelExporter(
            "tests", "TestCustomer", export_settings=export_settings
        )
        with self.assertRaises(ExportError):
            exporter.validate_filters({"email_client__icontains": "alpha"})

    def test_normalize_ordering_allows_only_allowlisted(self):
        export_settings = self._settings(["nom_client"])
        exporter = ModelExporter(
            "tests", "TestCustomer", export_settings=export_settings
        )
        self.assertEqual(exporter._normalize_ordering(["nom_client"]), ["nom_client"])
        with self.assertRaises(ExportError):
            exporter._normalize_ordering(["email_client"])

    def test_export_to_excel_returns_bytes(self):
        if not EXCEL_AVAILABLE:
            self.skipTest("openpyxl not available")
        export_settings = self._settings(["nom_client"])
        exporter = ModelExporter(
            "tests", "TestCustomer", export_settings=export_settings
        )
        data = exporter.export_to_excel(["nom_client"], ordering=["nom_client"])
        self.assertTrue(data.startswith(b"PK"))

    def test_export_to_csv_resolves_camel_case_and_preserves_falsy_values(self):
        TestCompany.objects.create(
            nom_entreprise="Acme",
            secteur_activite="Retail",
            adresse_entreprise="Main street",
            email_entreprise="acme@example.com",
            nombre_employes=0,
            est_active=False,
        )
        exporter = ModelExporter(
            "tests",
            "TestCompany",
            export_settings={
                "require_export_fields": False,
                "require_field_permissions": False,
            },
        )

        csv_data = exporter.export_to_csv(["nombreEmployes", "estActive"])
        rows = list(csv.reader(io.StringIO(csv_data)))
        self.assertEqual(rows[1][0], "0")
        self.assertEqual(rows[1][1], "❌")

    def test_validate_fields_rejects_unknown_accessor(self):
        exporter = ModelExporter(
            "tests",
            "TestCustomer",
            export_settings={
                "require_export_fields": False,
                "require_field_permissions": False,
            },
        )
        with self.assertRaises(ExportError):
            exporter.validate_fields(["doesNotExist"])

    def test_export_to_csv_accepts_camel_case_where_filters(self):
        export_settings = self._settings(
            ["nom_client", "est_actif"],
            filterable_fields={
                "tests.testcustomer": ["nom_client", "est_actif"],
            },
        )
        exporter = ModelExporter(
            "tests", "TestCustomer", export_settings=export_settings
        )

        csv_data = exporter.export_to_csv(
            ["nom_client", "est_actif"],
            variables={
                "where": {
                    "AND": [
                        {"estActif": {"eq": True}},
                        {"nomClient": {"icontains": "alp"}},
                    ]
                }
            },
        )

        rows = list(csv.reader(io.StringIO(csv_data)))
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][0], "Alpha")
        self.assertNotEqual(rows[1][1], "âŒ")

    def test_validate_filter_input_normalizes_camel_case_fields(self):
        export_settings = self._settings(
            ["nom_client", "est_actif"],
            filterable_fields={"tests.testcustomer": ["nom_client", "est_actif"]},
        )
        exporter = ModelExporter(
            "tests", "TestCustomer", export_settings=export_settings
        )

        exporter.validate_filter_input({"estActif": {"eq": True}})

        with self.assertRaises(ExportError) as context:
            exporter.validate_filter_input({"emailClient": {"icontains": "@"}})
        self.assertIn("email_client", str(context.exception))

    def test_validate_group_by_normalizes_camel_case(self):
        export_settings = self._settings(["nom_client", "est_actif"])
        exporter = ModelExporter(
            "tests", "TestCustomer", export_settings=export_settings
        )
        self.assertEqual(exporter.validate_group_by("estActif"), "est_actif")

    def test_validate_group_by_rejects_unallowlisted_field(self):
        export_settings = self._settings(["nom_client"])
        exporter = ModelExporter(
            "tests", "TestCustomer", export_settings=export_settings
        )
        with self.assertRaises(ExportError):
            exporter.validate_group_by("emailClient")

    def test_export_to_excel_with_group_by_returns_valid_workbook(self):
        if not EXCEL_AVAILABLE:
            self.skipTest("openpyxl not available")

        import openpyxl

        export_settings = self._settings(["nom_client", "est_actif"])
        exporter = ModelExporter(
            "tests", "TestCustomer", export_settings=export_settings
        )

        payload = exporter.export_to_excel(
            ["nom_client", "est_actif"],
            ordering=["nom_client"],
            group_by="estActif",
        )
        workbook = openpyxl.load_workbook(io.BytesIO(payload))
        sheet = workbook.active

        self.assertEqual(sheet.cell(row=1, column=1).value, "#")
        self.assertIn("Est", str(sheet.cell(row=2, column=1).value))
        merged_ranges = {str(cell_range) for cell_range in sheet.merged_cells.ranges}
        self.assertIn("A2:C2", merged_ranges)

