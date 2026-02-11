import pytest
from csv import reader
from io import BytesIO
from django.contrib.auth import get_user_model
from django.test import RequestFactory

from rail_django.extensions.auth import JWTManager
from rail_django.extensions.excel.builder import OPENPYXL_AVAILABLE
from rail_django.extensions.importing.views import ModelImportTemplateDownloadView
from test_app.models import Category

try:  # pragma: no cover - optional dependency
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    load_workbook = None
    get_column_letter = None

pytestmark = [pytest.mark.integration, pytest.mark.django_db]


def _bearer_token_for_user(user) -> str:
    return JWTManager.generate_token(user)["token"]


def test_model_import_template_download_returns_csv():
    user = get_user_model().objects.create_superuser(
        username="import_template_download_admin",
        email="import_template_download_admin@example.com",
        password="pass",
    )
    token = _bearer_token_for_user(user)

    request = RequestFactory().get(
        "/api/v1/import/templates/test_app/product/",
        {"format": "csv"},
        HTTP_AUTHORIZATION=f"Bearer {token}",
    )
    response = ModelImportTemplateDownloadView.as_view()(
        request,
        app_label="test_app",
        model_name="product",
    )

    assert response.status_code == 200
    assert response["Content-Disposition"] == (
        'attachment; filename="test_app-product-import-template.csv"'
    )

    lines = list(reader(response.content.decode("utf-8-sig").splitlines()))
    assert "name" in lines[0]
    assert "price" in lines[0]
    # CSV includes one prefilled default row.
    assert len(lines) == 2


def test_model_import_template_download_honors_selected_fields_and_keeps_required_columns():
    user = get_user_model().objects.create_superuser(
        username="import_template_download_selected_fields_admin",
        email="import_template_download_selected_fields_admin@example.com",
        password="pass",
    )
    token = _bearer_token_for_user(user)

    request = RequestFactory().get(
        "/api/v1/import/templates/test_app/product/",
        {
            "format": "csv",
            "fields": ["inventory_count"],
        },
        HTTP_AUTHORIZATION=f"Bearer {token}",
    )
    response = ModelImportTemplateDownloadView.as_view()(
        request,
        app_label="test_app",
        model_name="product",
    )

    assert response.status_code == 200
    lines = list(reader(response.content.decode("utf-8-sig").splitlines()))
    headers = lines[0]
    assert headers == ["name", "price", "inventory_count"]
    # inventory_count has django default=0 and should be prefilled.
    assert lines[1][2] == "0"


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl is required for xlsx rendering")
def test_model_import_template_download_returns_xlsx():
    Category.objects.create(name="Hardware", description="Physical goods")
    Category.objects.create(name="Software", description="Digital goods")

    user = get_user_model().objects.create_superuser(
        username="import_template_download_admin_xlsx",
        email="import_template_download_admin_xlsx@example.com",
        password="pass",
    )
    token = _bearer_token_for_user(user)

    request = RequestFactory().get(
        "/api/v1/import/templates/test_app/product/",
        {"format": "xlsx"},
        HTTP_AUTHORIZATION=f"Bearer {token}",
    )
    response = ModelImportTemplateDownloadView.as_view()(
        request,
        app_label="test_app",
        model_name="product",
    )

    assert response.status_code == 200
    assert (
        response["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response["Content-Disposition"] == (
        'attachment; filename="test_app-product-import-template.xlsx"'
    )
    # XLSX files are ZIP containers and begin with PK magic bytes.
    assert response.content.startswith(b"PK")

    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    assert "name" in headers
    assert "price" in headers
    assert "inventory_count" in headers
    # Header + 500 prefilled template rows.
    assert sheet.max_row == 501

    inventory_col = headers.index("inventory_count") + 1
    assert sheet.cell(row=2, column=inventory_col).value == 0

    category_col = headers.index("category") + 1
    category_letter = get_column_letter(category_col)
    validations = list(sheet.data_validations.dataValidation)
    assert any(
        str(validation.formula1).startswith("'_choices'!")
        and f"{category_letter}2:{category_letter}501" in str(validation.sqref)
        for validation in validations
    )
    assert "_choices" in workbook.sheetnames
    assert workbook["_choices"].sheet_state == "hidden"


def test_model_import_template_download_requires_authentication():
    request = RequestFactory().get(
        "/api/v1/import/templates/test_app/product/",
        {"format": "csv"},
    )
    response = ModelImportTemplateDownloadView.as_view()(
        request,
        app_label="test_app",
        model_name="product",
    )

    assert response.status_code == 401
